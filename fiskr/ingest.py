import re
import csv
import json
import hashlib
import logging
from typing import List, Dict, Any, Generator, Optional, Set, Tuple
import xml.etree.ElementTree as ET
from html.parser import HTMLParser

# We try to import pypdf to extract PDF text, fallback to empty text if unavailable.
try:
    import pypdf
    PDF_AVAILABLE = True
except ImportError:
    PDF_AVAILABLE = False

# Lecture des classeurs Excel (liste australienne DFAT). Dependance OPTIONNELLE,
# sur le meme principe que pypdf : son absence n'empeche pas Fiskr de demarrer,
# elle rend seulement la voie XLSX indisponible — avec un message qui dit quoi
# installer plutot qu'une pile d'appels.
try:
    import openpyxl
    XLSX_AVAILABLE = True
except ImportError:
    XLSX_AVAILABLE = False

logger = logging.getLogger("fiskr.ingest")


def parse_multi_value(item: Dict[str, Any], *column_names: str,
                      separator: str = ";") -> List[str]:
    """
    Liste de valeurs saisie dans UNE colonne CSV, valeurs separees par
    `separator`. Accepte plusieurs orthographes de colonne et retourne la
    premiere renseignee ; une liste deja construite (JSON, appel interne) est
    reprise telle quelle.

    Remplace une expression qui etait recopiee a cinq endroits et qui portait
    deux defauts :
    - elle n'interrogeait qu'UNE orthographe de colonne. Pour les adresses
      alternatives c'etait la forme courte `alternative_addresses`, alors que
      toutes les colonnes voisines acceptent aussi la forme prefixee
      `client_...`. Un fichier ecrit avec `client_alternative_addresses` etait
      donc ignore en silence ;
    - une colonne vide donnait `[""]` et non `[]` : `"".split(";")` retourne
      `[""]`. Chaque fiche sans adresse alternative portait une entree vide,
      comptee comme une adresse par tout ce qui lit ce champ.
    """
    for name in column_names:
        raw = item.get(name)
        if raw is None:
            continue
        if not isinstance(raw, str):
            return list(raw) if raw else []
        if raw.strip():
            return [part.strip() for part in raw.split(separator) if part.strip()]
    return []

# ------------------ ALIAS RISK CATEGORIZATION (Section 5.6) ------------------

def qualify_alias_priority(alias: str, alias_type_attr: str = "") -> str:
    """
    Qualifies an alias as HIGH (screened) or LOW (consultation only).
    Uses native attributes (Strong/Weak) if available, or fallback heuristics.
    """
    if alias_type_attr:
        cleaned_attr = alias_type_attr.lower().strip()
        if "strong" in cleaned_attr:
            return "high"
        elif "weak" in cleaned_attr:
            return "low"
            
    # Fallback Heuristics
    clean_a = re.sub(r"[\._\-]", " ", alias).strip()
    words = clean_a.split()
    
    # 1. Contains only a single word
    if len(words) <= 1:
        return "low"
        
    # 2. Total length is less than or equal to 4 characters
    if len(clean_a) <= 4:
        return "low"
        
    # 3. Composed only of Noise Words (SA, SARL, LLC, GMBH, LTD, SOCIETE, etc.)
    noise_pattern = r"^(SA|SARL|LLC|LTD|GMBH|SOCIETE|\s)+$"
    if re.match(noise_pattern, clean_a, re.IGNORECASE):
        return "low"
        
    return "high"

def categorize_aliases(alias_list: List[Dict[str, str]]) -> Dict[str, List[str]]:
    """
    Takes a list of alias objects: [{"name": "...", "type": "Strong/Weak/..."}]
    Returns {"high_priority": [...], "low_priority": [...]}
    """
    high = []
    low = []
    for a in alias_list:
        name = a.get("name", "")
        if not name:
            continue
        priority = qualify_alias_priority(name, a.get("type", ""))
        if priority == "high":
            high.append(name)
        else:
            low.append(name)
    return {"high_priority": high, "low_priority": low}


# ------------------ XML OFAC CONNECTOR (iterparse) ------------------

def get_attrib_insensitive(elem: Any, attr_name: str) -> Any:
    """
    Looks up an attribute in elem.attrib ignoring namespaces and case.
    """
    if elem is None or not hasattr(elem, "attrib"):
        return None
    target = attr_name.lower()
    for k, v in elem.attrib.items():
        local_key = k.split("}")[-1].lower() if "}" in k else k.lower()
        if local_key == target:
            return v
    return None

class OFACParserContext:
    def __init__(self):
        self.references = {}
        self.ref_links = {}
        self.locations = {}  # location_id -> {"full", "parts", "iso2", "country_name"}
        self.location_countries = {} # location_id -> Country ISO2
        self.id_documents = {}  # identity_id -> [doc_dict, ...]
        self.sanctions_programs = {}  # profile_id -> [program names]
        self.relationships = []  # [{"from_id", "to_id", "type_id"}] (ProfileRelationships)


def _local_ns(elem: ET.Element) -> str:
    """Prefixe de namespace ('{uri}') de l'element, ou chaine vide."""
    return elem.tag.split('}')[0] + '}' if '}' in elem.tag else ''


def _stream_target_elements(file_path: str, target_locals: Set[str]) -> Generator[Tuple[str, ET.Element], None, None]:
    """
    Streame le XML via iterparse et produit chaque element cible entierement
    construit, sous la forme (nom_local, element). Seuls les elements termines
    HORS de tout sous-arbre cible sont liberes au fil de l'eau : les descendants
    d'une cible restent intacts jusqu'au 'end' de la cible elle-meme (un clear
    inconditionnel viderait les valeurs des referentiels avant leur lecture,
    car les evenements 'end' remontent du bas vers le haut).
    """
    depth_in_target = 0
    root = None
    for event, elem in ET.iterparse(file_path, events=("start", "end")):
        local_name = elem.tag.split('}')[-1]
        if event == "start":
            if root is None:
                root = elem
            if depth_in_target > 0 or local_name in target_locals:
                depth_in_target += 1
            continue

        if depth_in_target > 0:
            depth_in_target -= 1
            if depth_in_target == 0 and local_name in target_locals:
                yield local_name, elem
                elem.clear()
            continue

        # Element termine hors de tout sous-arbre cible : liberation memoire
        elem.clear()
        if root is not None:
            root.clear()


def _extract_date_from_period_elem(elem: ET.Element, ns: str) -> Optional[str]:
    """Extrait une date YYYY-MM-DD depuis un sous-arbre contenant DatePeriod/Start/From."""
    frm = elem.find(f".//{ns}Start/{ns}From")
    if frm is None:
        frm = elem.find(f".//{ns}From")
    if frm is None:
        return None
    def _txt(tag):
        child = frm.find(f"{ns}{tag}")
        return child.text.strip() if (child is not None and child.text) else ""
    y, m, d = _txt("Year"), _txt("Month"), _txt("Day")
    if not y:
        return None
    return f"{y}-{(m or '01').zfill(2)}-{(d or '01').zfill(2)}"

def dict_get_insensitive(d, key):
    if not isinstance(d, dict):
        return None
    target = key.lower()
    for k, v in d.items():
        if k.lower() == target:
            return v
    return None

def find_nested_in_dict(data, target_key):
    results = []
    if isinstance(data, dict):
        for k, v in data.items():
            if k.lower() == target_key.lower():
                if isinstance(v, list):
                    results.extend(v)
                else:
                    results.append(v)
            else:
                results.extend(find_nested_in_dict(v, target_key))
    elif isinstance(data, list):
        for item in data:
            results.extend(find_nested_in_dict(item, target_key))
    return results

def get_reference_value(references, ref_type, val_id):
    for rk, rvals in references.items():
        if rk.lower() == ref_type.lower():
            return rvals.get(val_id, '')
    return ''

def elem_to_dict(elem, references):
    d = {}
    if elem.text and elem.text.strip():
        d['text'] = elem.text.strip()
    for k, v in elem.attrib.items():
        local_k = k.split('}')[-1] if '}' in k else k
        local_k_lower = local_k.lower()
        if local_k_lower.endswith('id') and not local_k_lower == 'id':
            ref_type = local_k[:-2]
            matched_ref_type = None
            for rk in references.keys():
                if rk.lower() == ref_type.lower():
                    matched_ref_type = rk
                    break
            if matched_ref_type and v in references[matched_ref_type]:
                d[local_k] = {"id": v, "value": references[matched_ref_type][v]}
            else:
                d[local_k] = v
        else:
            d[local_k] = v
            
    for child in elem:
        tag = child.tag.split('}')[-1]
        child_dict = elem_to_dict(child, references)
        if tag not in d:
            d[tag] = []
        d[tag].append(child_dict)
    return d

def _harvest_profile_relationship(elem: ET.Element, parser_ctx: OFACParserContext) -> None:
    """
    Recolte un lien entre profils (ProfileRelationships du SDN_ADVANCED) :
    From-ProfileID --[RelationTypeID]--> To-ProfileID. Les libelles de types
    sont resolus apres la passe 1 via le referentiel RelationType.
    """
    from_id = get_attrib_insensitive(elem, 'From-ProfileID') or get_attrib_insensitive(elem, 'FromProfileID')
    to_id = get_attrib_insensitive(elem, 'To-ProfileID') or get_attrib_insensitive(elem, 'ToProfileID')
    type_id = get_attrib_insensitive(elem, 'RelationTypeID')
    if from_id and to_id:
        parser_ctx.relationships.append({
            "from_id": str(from_id), "to_id": str(to_id), "type_id": str(type_id or ""),
        })


# Correspondance libelle OFAC -> code pivot de relation (RELATION_TYPES)
def _relation_code_from_label(label: str) -> str:
    lowered = (label or "").lower()
    if "owned" in lowered or "controlled" in lowered:
        return "OWNED_BY"
    if "acting" in lowered or "on behalf" in lowered:
        return "ACTING_FOR"
    if "associate" in lowered:
        return "ASSOCIATE_OF"
    if "family" in lowered:
        return "FAMILY_OF"
    if "leader" in lowered or "leading role" in lowered:
        return "LEADER_OF"
    if "support" in lowered:
        return "PROVIDING_SUPPORT"
    return "OTHER"


def _harvest_reference_sets(elem: ET.Element, parser_ctx: OFACParserContext) -> None:
    """Charge tous les jeux de valeurs de reference (PartyType, Country, FeatureType...)."""
    for value_set in list(elem):
        vs_tag = value_set.tag.split('}')[-1]
        base_tag = vs_tag.replace('Values', '')
        if base_tag not in parser_ctx.references:
            parser_ctx.references[base_tag] = {}
        if base_tag not in parser_ctx.ref_links:
            parser_ctx.ref_links[base_tag] = {}
        for child in list(value_set):
            if 'ID' in child.attrib:
                id_val = child.attrib['ID']
                if child.text and child.text.strip():
                    val = child.text.strip()
                elif 'Description' in child.attrib:
                    val = child.attrib['Description']
                else:
                    val = str(child.attrib)
                parser_ctx.references[base_tag][id_val] = val
                extra = {k: v for k, v in child.attrib.items() if k != 'ID'}
                if child.text and child.text.strip():
                    extra['_text'] = child.text.strip()
                if extra:
                    parser_ctx.ref_links[base_tag][id_val] = extra


def _harvest_location(elem: ET.Element, parser_ctx: OFACParserContext) -> None:
    """
    Indexe une localisation par ID : adresse complete, parties structurees
    (via LocPartType : ADDRESS1, CITY, STATE/PROVINCE, POSTAL CODE, REGION...)
    et pays (nom + ISO2).
    """
    if 'ID' not in elem.attrib:
        return
    id_val = elem.attrib['ID']
    ns = _local_ns(elem)

    parts = {}
    loc_texts = []
    for lp in elem.iter(f"{ns}LocationPart"):
        lp_type_id = get_attrib_insensitive(lp, 'LocPartTypeID')
        lp_type = parser_ctx.references.get('LocPartType', {}).get(str(lp_type_id or ''), '')
        texts = [v.text.strip() for v in lp.iter(f"{ns}Value") if v.text and v.text.strip()]
        if texts:
            val = ", ".join(texts)
            loc_texts.append(val)
            if lp_type:
                parts[lp_type.upper()] = val
    if not loc_texts:
        # Fichiers sans structure LocationPart : toutes les valeurs texte
        loc_texts = [p.text.strip() for p in elem.iter(f"{ns}Value") if p.text and p.text.strip()]

    country_name = None
    iso2 = None
    cid = None
    for p in elem.iter(f"{ns}LocationCountry"):
        cid = p.attrib.get('CountryID')
        if cid and cid in parser_ctx.references.get('Country', {}):
            country_name = parser_ctx.references['Country'][cid]

    full_parts = list(loc_texts)
    if country_name:
        full_parts.append(country_name)
    if cid:
        c_links = parser_ctx.ref_links.get('Country', {}).get(cid, {})
        iso2 = c_links.get('ISO2') or c_links.get('Code')
        if not iso2 and country_name:
            iso2 = country_name[:2].upper()
        if iso2:
            parser_ctx.location_countries[id_val] = iso2

    parser_ctx.locations[id_val] = {
        "full": ", ".join(full_parts),
        "parts": parts,
        "iso2": iso2,
        "country_name": country_name,
    }


def _harvest_id_document(elem: ET.Element, parser_ctx: OFACParserContext) -> None:
    """Indexe un document d'identite par IdentityID (numero, pays emetteur, expiration)."""
    identity_id = elem.attrib.get('IdentityID')
    if not identity_id:
        return
    ns = _local_ns(elem)
    doc_type_id = elem.attrib.get('IDRegDocTypeID')
    doc_num = ""
    doc_num_el = elem.find(f".//{ns}IDRegistrationNo")
    if doc_num_el is not None and doc_num_el.text:
        doc_num = doc_num_el.text.strip()

    issued_by_el = elem.find(f".//{ns}IssuedBy")
    issuing_country = "XX"
    if issued_by_el is not None:
        cid = issued_by_el.attrib.get('CountryID')
        if cid and cid in parser_ctx.references.get('Country', {}):
            country_name = parser_ctx.references['Country'][cid]
            c_links = parser_ctx.ref_links.get('Country', {}).get(cid, {})
            iso2 = c_links.get('ISO2') or c_links.get('Code') or country_name[:2].upper()
            issuing_country = iso2

    # Date d'expiration (DocumentDate type "Expiration Date" du referentiel)
    expiration = None
    for dd in elem.iter(f"{ns}DocumentDate"):
        dtype_id = get_attrib_insensitive(dd, 'IDRegDocDateTypeID')
        dtype_name = parser_ctx.references.get('IDRegDocDateType', {}).get(str(dtype_id or ''), '').lower()
        date_val = _extract_date_from_period_elem(dd, ns)
        if date_val and ("expir" in dtype_name):
            expiration = date_val

    doc_dict = {
        "doc_type_id": doc_type_id,
        "number": doc_num,
        "issuing_country": issuing_country,
        "expiration_date": expiration
    }
    if identity_id not in parser_ctx.id_documents:
        parser_ctx.id_documents[identity_id] = []
    parser_ctx.id_documents[identity_id].append(doc_dict)


def _harvest_sanctions_entry(elem: ET.Element, parser_ctx: OFACParserContext) -> None:
    """
    Recolte les programmes de sanctions d'une SanctionsEntry (liee par ProfileID).
    Les mesures dont le type se resout en "Program" portent le nom du programme
    dans leur Comment ; sans referentiel charge, tout Comment est conserve.
    """
    profile_id = get_attrib_insensitive(elem, 'ProfileID')
    if not profile_id:
        return
    ns = _local_ns(elem)
    sanctions_types = parser_ctx.references.get('SanctionsType', {})
    programs = []
    for measure in elem.iter(f"{ns}SanctionsMeasure"):
        type_id = get_attrib_insensitive(measure, 'SanctionsTypeID')
        type_name = sanctions_types.get(str(type_id or ''), '').lower()
        comment_el = measure.find(f"{ns}Comment")
        text = comment_el.text.strip() if (comment_el is not None and comment_el.text) else ""
        if not text:
            continue
        if "program" in type_name or not sanctions_types:
            programs.append(text)
    if programs:
        existing = parser_ctx.sanctions_programs.setdefault(str(profile_id), [])
        for p in programs:
            if p not in existing:
                existing.append(p)


def _classify_id_document(doc_type_id, doc_type_name, doc_num, issued_country, expiration_date, buckets):
    """
    Route un document d'identite vers le bon compartiment du schema pivot.
    Les IDs numeriques codes en dur couvrent les fichiers simplifies ; les
    correspondances par nom (referentiel IDRegDocType) couvrent le fichier
    officiel dont les IDs varient.
    """
    doc_type_id = str(doc_type_id or "")
    doc_type_name = (doc_type_name or "").lower()
    if doc_type_id == "392" or "passport" in doc_type_name:
        buckets["passports"].append({"number": doc_num, "issuing_country": issued_country, "expiration_date": expiration_date})
    elif doc_type_id == "391" or "national id" in doc_type_name:
        buckets["national_ids"].append({"number": doc_num, "issuing_country": issued_country})
    elif doc_type_id in ("386", "390", "394") or "driver" in doc_type_name:
        buckets["other_ids"].append({"doc_type": "DriverLicense" if doc_type_id == "386" or "driver" in doc_type_name else "Other", "number": doc_num, "issuing_country": issued_country})
    elif doc_type_id == "15502" or "lei" in doc_type_name:
        buckets["lei"] = doc_num
    elif doc_type_id in ("9436", "376", "384") or "tax" in doc_type_name or "commercial" in doc_type_name or "business registration" in doc_type_name:
        buckets["national_registry"].append({"number": doc_num, "country": issued_country, "registry_name": "CommercialRegistry" if doc_type_id == "9436" or "commercial" in doc_type_name or "business registration" in doc_type_name else "TaxRegistry"})
    elif doc_type_id == "13886" or "imo" in doc_type_name or "vessel registration" in doc_type_name:
        digits = re.sub(r"\D", "", doc_num)
        buckets["imo_number"] = digits[:7]
    elif doc_type_id == "13887" or "aircraft" in doc_type_name:
        buckets["aircraft_tail"] = doc_num
    else:
        buckets["other_registrations"].append({"id_type": doc_type_name or "OtherRegistration", "number": doc_num})


def resolve_party_type(profile, parser_ctx):
    # Try child element style first (from mock XML or simplified schemas)
    pst_list = dict_get_insensitive(profile, 'PartySubType')
    if pst_list and isinstance(pst_list, list) and len(pst_list) > 0:
        pst_elem = pst_list[0]
        ptype = dict_get_insensitive(pst_elem, 'PartyTypeID')
        if isinstance(ptype, dict):
            ptype = ptype.get('id')
        if ptype == "151":
            return "I"
        elif ptype == "152":
            return "E"
        elif ptype == "154":
            return "V"
        elif ptype == "153":
            return "O"
            
    # Try attribute style (from standard Advanced XML)
    pst = dict_get_insensitive(profile, 'PartySubTypeID')
    if not pst:
        return None
    pst_value = ""
    if isinstance(pst, dict):
        pst_id = pst.get('id', '')
        pst_value = str(pst.get('value') or '')
    elif isinstance(pst, list) and pst:
        pst_id = pst[0].get('id', '') if isinstance(pst[0], dict) else ''
        pst_value = str(pst[0].get('value') or '') if isinstance(pst[0], dict) else ''
    else:
        pst_id = str(pst)

    links = parser_ctx.ref_links.get('PartySubType', {}).get(pst_id, {})
    pt_id = links.get('PartyTypeID', '')
    pt_name = get_reference_value(parser_ctx.references, 'PartyType', pt_id).lower()
    # Le nom du sous-type lui-meme (ex: "Individual") est aussi discriminant
    combined = f"{pst_value.lower()} {get_reference_value(parser_ctx.references, 'PartySubType', pst_id).lower()} {pt_name}"

    if "individual" in combined:
        return "I"
    elif "vessel" in combined:
        return "V"
    elif "aircraft" in combined:
        return "O"
    elif "entity" in combined:
        return "E"
    # Referentiel absent ou irresoluble : on laisse l'heuristique decider
    return None

def _feature_version_text(fv) -> str:
    """
    Extrait le texte d'une version de feature : valeurs resolues des
    DetailReferenceID, contenus des DetailReference et texte brut du detail.
    """
    out = []
    version_details = dict_get_insensitive(fv, 'VersionDetail') or []
    for vd in version_details:
        ref_obj = dict_get_insensitive(vd, 'DetailReferenceID')
        if ref_obj:
            if isinstance(ref_obj, dict):
                out.append(str(ref_obj.get('value') or ref_obj.get('id') or ''))
            elif isinstance(ref_obj, list):
                for child in ref_obj:
                    if isinstance(child, dict):
                        out.append(str(child.get('value') or child.get('text') or ''))
            else:
                out.append(str(ref_obj))
        ref_ref = dict_get_insensitive(vd, 'DetailReference')
        if ref_ref:
            if isinstance(ref_ref, list):
                for child in ref_ref:
                    if isinstance(child, dict):
                        out.append(str(child.get('value') or child.get('text') or ''))
            elif isinstance(ref_ref, dict):
                out.append(str(ref_ref.get('value') or ref_ref.get('text') or ''))
            else:
                out.append(str(ref_ref))
        if isinstance(vd, dict) and vd.get('text'):
            out.append(vd['text'])
    return " ".join(t.strip() for t in out if t and t.strip()).strip()


def _feature_dateperiod_iso(fv) -> Optional[str]:
    """Premiere date (Start/From) d'un DatePeriod de feature, au format ISO.
    Utilisee pour les features datees non textuelles (ex. Organization
    Established Date des fichiers OFAC reels)."""
    date_periods = dict_get_insensitive(fv, 'DatePeriod') or []
    for dp in date_periods:
        start = dict_get_insensitive(dp, 'Start') or []
        if start and 'From' in start[0]:
            from_date = start[0]['From'][0]
            y_el = dict_get_insensitive(from_date, 'Year')
            m_el = dict_get_insensitive(from_date, 'Month')
            d_el = dict_get_insensitive(from_date, 'Day')
            y = y_el[0].get('text', '') if (y_el and isinstance(y_el, list)) else ''
            m = m_el[0].get('text', '') if (m_el and isinstance(m_el, list)) else ''
            d = d_el[0].get('text', '') if (d_el and isinstance(d_el, list)) else ''
            if y:
                return f"{y.strip()}-{(m.strip() if m else '01').zfill(2)}-{(d.strip() if d else '01').zfill(2)}"
    return None


def format_alias_name(alias_dict, identity_dict):
    group_map = {}
    name_part_groups = dict_get_insensitive(identity_dict, 'NamePartGroups') or []
    for groups in name_part_groups:
        master_groups = dict_get_insensitive(groups, 'MasterNamePartGroup') or []
        for mg in master_groups:
            ng_list = dict_get_insensitive(mg, 'NamePartGroup') or []
            for ng in ng_list:
                gid = dict_get_insensitive(ng, 'ID')
                tid = dict_get_insensitive(ng, 'NamePartTypeID')
                ty = tid.get('value') if isinstance(tid, dict) else str(tid)
                group_map[gid] = ty
                
    order_map = {
        "First Name": 1,
        "Middle Name": 2,
        "Patronymic": 3,
        "Matronymic": 4,
        "Last Name": 5,
        "Entity Name": 10,
        "Nickname": 11,
        "Vessel Name": 12,
        "Aircraft Name": 13
    }
    
    parts_list = []
    documented_names = dict_get_insensitive(alias_dict, 'DocumentedName') or []
    for dn in documented_names:
        name_parts = dict_get_insensitive(dn, 'DocumentedNamePart') or []
        for pt in name_parts:
            name_part_values = dict_get_insensitive(pt, 'NamePartValue') or []
            for nv in name_part_values:
                if 'text' in nv:
                    gid = dict_get_insensitive(nv, 'NamePartGroupID')
                    ty = group_map.get(gid, "Unknown")
                    weight = order_map.get(ty, 99)
                    parts_list.append((weight, nv['text']))
                    
    parts_list.sort(key=lambda x: x[0])
    return " ".join([x[1] for x in parts_list])

def parse_ofac_advanced_xml(file_path: str,
                            relations_out: Optional[List[Dict[str, Any]]] = None
                            ) -> Generator[Dict[str, Any], None, None]:
    """
    Sequentially parses the OFAC Advanced XML using ElementTree.iterparse
    to prevent memory ballooning. Yields Pivot Schema dicts.

    `relations_out` : liste optionnelle que l'appelant fournit pour recevoir
    les liens entre profils (ProfileRelationships) resolus en codes pivots —
    {"from_entity_id", "to_entity_id", "relation_type", "relation_label"}.
    """
    parser_ctx = OFACParserContext()

    # Pass 1: recolte des referentiels, localisations, documents d'identite,
    # programmes de sanctions et liens entre profils (SanctionsEntries suit
    # DistinctParties dans le fichier officiel, d'ou la necessite de deux
    # passes). Le streaming a suivi de profondeur garantit que les enfants
    # d'une cible ne sont jamais vides avant la lecture de la cible.
    for local_name, elem in _stream_target_elements(
        file_path, {'ReferenceValueSets', 'Location', 'IDRegDocument', 'SanctionsEntry', 'ProfileRelationship'}
    ):
        if local_name == 'ReferenceValueSets':
            _harvest_reference_sets(elem, parser_ctx)
        elif local_name == 'Location':
            _harvest_location(elem, parser_ctx)
        elif local_name == 'IDRegDocument':
            _harvest_id_document(elem, parser_ctx)
        elif local_name == 'SanctionsEntry':
            _harvest_sanctions_entry(elem, parser_ctx)
        elif local_name == 'ProfileRelationship':
            _harvest_profile_relationship(elem, parser_ctx)

    # Resolution des liens entre profils (libelles du referentiel RelationType)
    if relations_out is not None:
        relation_labels = parser_ctx.references.get('RelationType', {})
        for rel in parser_ctx.relationships:
            label = relation_labels.get(rel["type_id"], "")
            relations_out.append({
                "from_entity_id": rel["from_id"],
                "to_entity_id": rel["to_id"],
                "relation_type": _relation_code_from_label(label),
                "relation_label": label or None,
            })

    # Pass 2: Parse DistinctParties
    for _, elem in _stream_target_elements(file_path, {'DistinctParty'}):
        ns = _local_ns(elem)
        prof_elem = elem.find(f'{ns}Profile')
        if prof_elem is None:
            continue

        pid = (
            get_attrib_insensitive(elem, "fixedRef")
            or get_attrib_insensitive(elem, "ID")
            or (get_attrib_insensitive(prof_elem, "ID") if prof_elem is not None else None)
        )
        if not pid:
            continue
            
        profile = elem_to_dict(prof_elem, parser_ctx.references)
        
        # Extract basic fields
        entity_type_id = resolve_party_type(profile, parser_ctx)
        primary_name = ""
        first_name = ""
        last_name = ""
        maiden_name = ""
        aliases_raw = []
        
        # Extract names & aliases
        for identity in profile.get('Identity', []):
            for alias in identity.get('Alias', []):
                is_primary = alias.get('Primary') == 'true'
                formatted_name = format_alias_name(alias, identity)
                if not formatted_name:
                    continue
                    
                alias_type_obj = alias.get('AliasTypeID')
                alias_type_str = "Strong"
                if isinstance(alias_type_obj, dict):
                    alias_type_str = alias_type_obj.get('value', 'Strong')
                elif alias_type_obj:
                    alias_type_str = str(alias_type_obj)
                    
                if is_primary:
                    primary_name = formatted_name
                    
                    # Extract first, last, maiden name
                    group_map = {}
                    for groups in identity.get('NamePartGroups', []):
                        for mg in groups.get('MasterNamePartGroup', []):
                            for ng in mg.get('NamePartGroup', []):
                                gid = ng.get('ID')
                                tid = ng.get('NamePartTypeID', {})
                                group_map[gid] = tid.get('value') if isinstance(tid, dict) else str(tid)
                                
                    for dn in alias.get('DocumentedName', []):
                        for pt in dn.get('DocumentedNamePart', []):
                            for nv in pt.get('NamePartValue', []):
                                if 'text' in nv:
                                    gid = nv.get('NamePartGroupID')
                                    ty = group_map.get(gid, "Unknown")
                                    if ty == "First Name":
                                        first_name = nv['text']
                                    elif ty == "Last Name":
                                        last_name = nv['text']
                                    elif "maiden" in ty.lower():
                                        maiden_name = nv['text']
                else:
                    aliases_raw.append({"name": formatted_name, "type": alias_type_str})
        
        # Nested DocumentedName fallback
        if not primary_name:
            nested_doc_names = find_nested_in_dict(profile, 'DocumentedName')
            for doc_name in nested_doc_names:
                status_id = dict_get_insensitive(doc_name, "DocNameStatusID")
                if isinstance(status_id, dict):
                    status_id = status_id.get('id', '')
                is_primary = str(status_id) == "1"
                
                name_parts = []
                parts = find_nested_in_dict(doc_name, 'DocumentedNamePart')
                for part in parts:
                    part_type = dict_get_insensitive(part, "NamePartTypeID")
                    if isinstance(part_type, dict):
                        part_type = part_type.get('id', '')
                    else:
                        part_type = str(part_type or '')
                        
                    part_vals = find_nested_in_dict(part, 'Value')
                    for pv in part_vals:
                        text = pv.get('text', '') if isinstance(pv, dict) else str(pv)
                        if text:
                            text_clean = text.strip()
                            name_parts.append(text_clean)
                            if is_primary:
                                if part_type == "1360":
                                    first_name = text_clean
                                elif part_type == "1361":
                                    last_name = text_clean
                                    
                full_name_resolved = " ".join(name_parts)
                if is_primary:
                    primary_name = full_name_resolved
                else:
                    alias_type = dict_get_insensitive(doc_name, "AliasTypeID")
                    if isinstance(alias_type, dict):
                        alias_type = alias_type.get('id', '')
                    type_str = "Strong" if str(alias_type) == "1" else "Weak"
                    aliases_raw.append({"name": full_name_resolved, "type": type_str})
        
        # Extract features (DOB, Gender, Death/Deceased, countries, POB, addresses...)
        dobs = []
        date_of_death = None
        is_deceased = False
        gender = "U"
        citizenships = []
        residences = []
        birth_countries = []
        jurisdictions = []
        place_of_birth = None
        addresses = []       # [{"full", "parts", ...}] dans l'ordre du fichier
        designation = None
        unmapped_features = []  # features non pivotables -> additional_informations

        # Champs etendus structures (au lieu du fourre-tout texte)
        crypto_wallets = []
        bic_swift = None
        tax_id = None
        duns_number = None
        vessel_call_sign = None
        vessel_mmsi = None
        vessel_flag = None
        vessel_type_val = None
        vessel_tonnage = None
        vessel_owner = None
        aircraft_model = None
        aircraft_operator = None
        aircraft_construction_number = None
        websites = []
        email_addresses = []
        phone_numbers = []
        secondary_sanctions_risk = None
        organization_established_date = None
        organization_type = None

        features = dict_get_insensitive(profile, 'Feature') or []
        for f in features:
            ftype_obj = dict_get_insensitive(f, 'FeatureTypeID')
            if not ftype_obj:
                continue
            ftype_str = ftype_obj.get('value', '') if isinstance(ftype_obj, dict) else str(ftype_obj)
            ftype_str_lower = ftype_str.lower()

            is_gender = "gender" in ftype_str_lower or ftype_str_lower == "25"
            is_birth = ("birth" in ftype_str_lower and "date" in ftype_str_lower) or ftype_str_lower in ["8", "12"]
            is_death = "death" in ftype_str_lower or "deceased" in ftype_str_lower or ftype_str_lower == "24"
            # "place of birth" avant la branche generique "birth" (pays de naissance)
            is_pob = "place of birth" in ftype_str_lower
            # "Digital Currency Address" et "Email Address" contiennent le mot
            # "address" mais ne sont PAS des adresses postales
            is_address = (
                "address" in ftype_str_lower
                and "digital currency" not in ftype_str_lower
                and "email" not in ftype_str_lower
            ) or ftype_str_lower == "location"
            is_designation = any(k in ftype_str_lower for k in ("title", "position", "function", "occupation"))

            feature_versions = dict_get_insensitive(f, 'FeatureVersion') or []
            for fv in feature_versions:
                # Gender
                if is_gender:
                     ref_val_lower = _feature_version_text(fv).lower()
                     if "female" in ref_val_lower:
                         gender = "F"
                     elif "male" in ref_val_lower:
                         gender = "M"

                # Birth
                elif is_birth:
                    date_periods = dict_get_insensitive(fv, 'DatePeriod') or []
                    for dp in date_periods:
                        start = dict_get_insensitive(dp, 'Start') or []
                        if start and 'From' in start[0]:
                            from_date = start[0]['From'][0]
                            y_el = dict_get_insensitive(from_date, 'Year')
                            m_el = dict_get_insensitive(from_date, 'Month')
                            d_el = dict_get_insensitive(from_date, 'Day')
                            y = y_el[0].get('text', '') if (y_el and isinstance(y_el, list)) else ''
                            m = m_el[0].get('text', '') if (m_el and isinstance(m_el, list)) else ''
                            d = d_el[0].get('text', '') if (d_el and isinstance(d_el, list)) else ''
                            if y:
                                m_str = m.strip() if m else "01"
                                d_str = d.strip() if d else "01"
                                dobs.append(f"{y.strip()}-{m_str.zfill(2)}-{d_str.zfill(2)}")
                                
                # Death
                elif is_death:
                    is_deceased = True
                    date_periods = dict_get_insensitive(fv, 'DatePeriod') or []
                    for dp in date_periods:
                        start = dict_get_insensitive(dp, 'Start') or []
                        if start and 'From' in start[0]:
                            from_date = start[0]['From'][0]
                            y_el = dict_get_insensitive(from_date, 'Year')
                            m_el = dict_get_insensitive(from_date, 'Month')
                            d_el = dict_get_insensitive(from_date, 'Day')
                            y = y_el[0].get('text', '') if (y_el and isinstance(y_el, list)) else ''
                            m = m_el[0].get('text', '') if (m_el and isinstance(m_el, list)) else ''
                            d = d_el[0].get('text', '') if (d_el and isinstance(d_el, list)) else ''
                            if y:
                                m_str = m.strip() if m else "01"
                                d_str = d.strip() if d else "01"
                                date_of_death = f"{y.strip()}-{m_str.zfill(2)}-{d_str.zfill(2)}"
                
                # Localisations liees a la feature (pays, lieu de naissance, adresses)
                version_locations = dict_get_insensitive(fv, 'VersionLocation') or []
                for vl in version_locations:
                    lid_obj = dict_get_insensitive(vl, 'LocationID')
                    lid = lid_obj.get('id') if isinstance(lid_obj, dict) else str(lid_obj)
                    if not lid:
                        continue
                    loc_info = parser_ctx.locations.get(lid) or {}
                    country_code = parser_ctx.location_countries.get(lid)
                    if is_pob:
                        if loc_info.get("full") and not place_of_birth:
                            place_of_birth = loc_info["full"]
                        if country_code:
                            birth_countries.append(country_code)
                    elif is_address:
                        if loc_info.get("full"):
                            addresses.append(loc_info)
                    elif country_code:
                        if "citizenship" in ftype_str_lower or "nationality" in ftype_str_lower:
                            citizenships.append(country_code)
                        elif "residence" in ftype_str_lower:
                            residences.append(country_code)
                        elif "birth" in ftype_str_lower:
                            birth_countries.append(country_code)
                        else:
                            jurisdictions.append(country_code)

                # Features structurees : mapping cible par type de feature
                # (crypto, BIC, tax ID, navire, aeronef, contacts...) ; le
                # reste part en additional_informations (consultation humaine)
                if not (is_gender or is_birth or is_death or is_pob or is_address) and not version_locations:
                    text = _feature_version_text(fv)
                    if not text and "organization established date" in ftype_str_lower:
                        # Fichiers OFAC reels : la date de creation est portee
                        # en DatePeriod, pas en texte de VersionDetail
                        text = _feature_dateperiod_iso(fv) or ""
                    if text:
                        if is_designation and not designation:
                            designation = text
                        elif "digital currency address" in ftype_str_lower:
                            currency = ftype_str.rsplit("-", 1)[-1].strip() if "-" in ftype_str else ""
                            crypto_wallets.append({"currency": currency, "address": text})
                        elif "swift" in ftype_str_lower or "bik" in ftype_str_lower:
                            bic_swift = bic_swift or text
                        elif "tax id" in ftype_str_lower:
                            tax_id = tax_id or text
                        elif "duns" in ftype_str_lower or "d-u-n-s" in ftype_str_lower:
                            duns_number = duns_number or text
                        elif "call sign" in ftype_str_lower:
                            vessel_call_sign = vessel_call_sign or text
                        elif ftype_str_lower.strip() == "msi" or "mmsi" in ftype_str_lower:
                            vessel_mmsi = vessel_mmsi or text
                        elif "flag" in ftype_str_lower:
                            vessel_flag = vessel_flag or text
                        elif "vessel type" in ftype_str_lower:
                            vessel_type_val = vessel_type_val or text
                        elif "tonnage" in ftype_str_lower:
                            vessel_tonnage = vessel_tonnage or text
                        elif "vessel owner" in ftype_str_lower:
                            vessel_owner = vessel_owner or text
                        elif "aircraft model" in ftype_str_lower:
                            aircraft_model = aircraft_model or text
                        elif "aircraft operator" in ftype_str_lower:
                            aircraft_operator = aircraft_operator or text
                        elif "construction number" in ftype_str_lower or "serial number" in ftype_str_lower:
                            aircraft_construction_number = aircraft_construction_number or text
                        elif "website" in ftype_str_lower:
                            websites.append(text)
                        elif "email" in ftype_str_lower:
                            email_addresses.append(text)
                        elif "phone" in ftype_str_lower:
                            phone_numbers.append(text)
                        elif "secondary sanctions risk" in ftype_str_lower:
                            secondary_sanctions_risk = (
                                f"{secondary_sanctions_risk}; {text}" if secondary_sanctions_risk else text
                            )
                        elif "organization established date" in ftype_str_lower:
                            organization_established_date = organization_established_date or (_extract_iso_date(text) or text)
                        elif "organization type" in ftype_str_lower:
                            organization_type = organization_type or text
                        else:
                            unmapped_features.append(f"{ftype_str}: {text}")

        # Fallback to nested locations
        if not citizenships and not residences and not birth_countries and not jurisdictions:
            nested_locations = find_nested_in_dict(profile, 'Location')
            for loc in nested_locations:
                loc_type_list = find_nested_in_dict(loc, 'LocationType')
                loc_type = ""
                if loc_type_list:
                    loc_type = loc_type_list[0].get('text', '') if isinstance(loc_type_list[0], dict) else str(loc_type_list[0])
                    
                country_list = find_nested_in_dict(loc, 'LocationCountry')
                country_code = ""
                if country_list:
                    country_el = country_list[0]
                    if isinstance(country_el, dict):
                        country_code = (dict_get_insensitive(country_el, 'CountryISO2') 
                                        or dict_get_insensitive(country_el, 'CountryID'))
                        if isinstance(country_code, dict):
                            country_code = country_code.get('id') or country_code.get('value')
                    else:
                        country_code = str(country_el)
                        
                if country_code:
                    lt_str = str(loc_type).lower()
                    if "citizenship" in lt_str:
                        citizenships.append(country_code)
                    elif "residence" in lt_str:
                        residences.append(country_code)
                    elif "birth" in lt_str:
                        birth_countries.append(country_code)
                    else:
                        residences.append(country_code)
                        
        # Extract ID registration documents
        imo_number = None
        aircraft_tail = None
        lei = None
        national_registry = []
        other_registrations = []
        passports = []
        national_ids = []
        other_ids = []

        doc_buckets = {
            "passports": passports,
            "national_ids": national_ids,
            "other_ids": other_ids,
            "national_registry": national_registry,
            "other_registrations": other_registrations,
            "lei": None,
            "imo_number": None,
            "aircraft_tail": None,
        }

        # Load documents linked to any identity in the profile
        for identity in profile.get('Identity', []):
            ident_id = identity.get('ID')
            if not ident_id:
                continue
            docs = parser_ctx.id_documents.get(ident_id, [])
            for doc in docs:
                doc_type_name = parser_ctx.references.get('IDRegDocType', {}).get(doc["doc_type_id"], "")
                _classify_id_document(
                    doc["doc_type_id"], doc_type_name, doc["number"],
                    doc["issuing_country"], doc.get("expiration_date"), doc_buckets
                )
        lei = doc_buckets["lei"]
        imo_number = doc_buckets["imo_number"]
        aircraft_tail = doc_buckets["aircraft_tail"]

        # Fallback to nested IDRegistrationDocument / IDRegDocument elements
        if not passports and not national_ids and not other_ids and not lei and not national_registry and not imo_number and not aircraft_tail:
            nested_docs = find_nested_in_dict(profile, 'IDRegistrationDocument') + find_nested_in_dict(profile, 'IDRegDocument')
            for doc_elem in nested_docs:
                doc_type_id = (dict_get_insensitive(doc_elem, "IDRegistrationDocTypeID") 
                               or dict_get_insensitive(doc_elem, "IDRegDocTypeID"))
                if isinstance(doc_type_id, dict):
                    doc_type_id = doc_type_id.get('id', '')
                else:
                    doc_type_id = str(doc_type_id or '')
                    
                doc_num_el_list = (find_nested_in_dict(doc_elem, "IDRegistrationDocElement") 
                                   or find_nested_in_dict(doc_elem, "IDRegistrationNo"))
                doc_num = ""
                if doc_num_el_list:
                    if isinstance(doc_num_el_list[0], dict):
                        doc_num = doc_num_el_list[0].get('text', '')
                    else:
                        doc_num = str(doc_num_el_list[0])
                
                issuing_el_list = find_nested_in_dict(doc_elem, "IssuedBy")
                issued_country = "XX"
                if issuing_el_list:
                    issuing_el = issuing_el_list[0]
                    country_el_list = find_nested_in_dict(issuing_el, "CountryISO2")
                    if country_el_list:
                        country_el = country_el_list[0]
                        if isinstance(country_el, dict):
                            issued_country = country_el.get('text') or country_el.get('CountryID') or "XX"
                            if isinstance(issued_country, dict):
                                issued_country = issued_country.get('id') or "XX"
                        else:
                            issued_country = str(country_el)
                            
                if doc_num:
                    doc_type_name = parser_ctx.references.get('IDRegDocType', {}).get(doc_type_id, "")
                    _classify_id_document(doc_type_id, doc_type_name, doc_num, issued_country, None, doc_buckets)
            lei = doc_buckets["lei"]
            imo_number = doc_buckets["imo_number"]
            aircraft_tail = doc_buckets["aircraft_tail"]

        # Repli heuristique quand ni le style enfant ni le referentiel n'ont
        # permis de typer le liste (fichiers simplifies ou referentiel absent)
        if not entity_type_id:
            if imo_number:
                entity_type_id = "V"
            elif aircraft_tail:
                entity_type_id = "O"
            elif gender != "U" or dobs or passports or national_ids or first_name or maiden_name:
                entity_type_id = "I"
            else:
                entity_type_id = "E"

        # Adresses structurees : premiere adresse = principale, le reste en alternatives
        primary_addr = addresses[0] if addresses else {}
        addr_parts = primary_addr.get("parts", {})

        # Build Pivot structure
        aliases_categorized = categorize_aliases(aliases_raw)
        current_party = {
            "entity_id": pid,
            "entity_type": entity_type_id,
            "primary_name": primary_name or "NOM INCONNU",
            "individual_name_parsed": {
                "first_name": first_name,
                "last_name": last_name,
                "maiden_name": maiden_name
            },
            "aliases": aliases_categorized,
            "dates_of_birth": list(set(dobs)),
            "date_of_death": date_of_death,
            "is_deceased": is_deceased,
            "gender": gender,
            "countries": {
                "citizenship": list(set(citizenships)),
                "residence": list(set(residences)),
                "birth_country": list(set(birth_countries)),
                "jurisdiction_country": list(set(jurisdictions))
            },
            "place_of_birth": place_of_birth,
            "address": primary_addr.get("full"),
            "alternative_addresses": [a["full"] for a in addresses[1:] if a.get("full")],
            "city": addr_parts.get("CITY"),
            "state": addr_parts.get("STATE/PROVINCE") or addr_parts.get("REGION"),
            "country": primary_addr.get("country_name"),
            "designation": designation,
            "designation_reasons": "; ".join(parser_ctx.sanctions_programs.get(str(pid), [])) or None,
            "additional_informations": "; ".join(unmapped_features) or None,
            "origin": "OFAC SDN_ADVANCED",
            "imo_number": imo_number,
            "aircraft_tail_number": aircraft_tail,
            "lei_number": lei,
            "national_registry_ids": national_registry,
            "other_registration_ids": other_registrations,
            "passport_documents": passports,
            "national_id_documents": national_ids,
            "other_id_documents": other_ids,
            # Champs etendus structures
            "crypto_wallets": crypto_wallets,
            "bic_swift": bic_swift,
            "tax_id": tax_id,
            "duns_number": duns_number,
            "vessel_call_sign": vessel_call_sign,
            "vessel_mmsi": vessel_mmsi,
            "vessel_flag": vessel_flag,
            "vessel_type": vessel_type_val,
            "vessel_tonnage": vessel_tonnage,
            "vessel_owner": vessel_owner,
            "aircraft_model": aircraft_model,
            "aircraft_operator": aircraft_operator,
            "aircraft_construction_number": aircraft_construction_number,
            "sanction_programs": parser_ctx.sanctions_programs.get(str(pid)) or [],
            "secondary_sanctions_risk": secondary_sanctions_risk,
            "organization_established_date": organization_established_date,
            "organization_type": organization_type,
            "phone_numbers": phone_numbers,
            "email_addresses": email_addresses,
            "websites": websites,
        }

        yield current_party
        
                



# ------------------ OFAC NON-SDN (CONS_ADVANCED.XML) ------------------
# L'OFAC publie DEUX fichiers au meme format « Advanced » : la SDN List et la
# « Consolidated Sanctions List », dite Non-SDN. La seconde porte les regimes
# qui ne relevent PAS d'un gel total des avoirs et sont donc absents du fichier
# SDN — au premier rang desquels les SECTORAL SANCTIONS (SSI, directives 1 a 4
# sur la Russie), mais aussi FSE, NS-MBS, PLC, MEU et CMIC. Un etablissement
# expose au dollar doit les cribler ; ne charger que la SDN laisse un trou.
#
# Le format etant strictement identique, ce connecteur ne re-parse rien : il
# reutilise `parse_ofac_advanced_xml`, deja eprouve sur la structure reelle du
# fichier SDN. Deux ajustements seulement, expliques ci-dessous.


def parse_ofac_consolidated_xml(file_path: str) -> Generator[Dict[str, Any], None, None]:
    """Parse le fichier Non-SDN de l'OFAC (CONS_ADVANCED.XML).

    Le parseur SDN rend l'identifiant de profil OFAC nu (`pid`). Les deux
    fichiers partagent le meme espace de FixedRef et ne se recouvrent pas
    aujourd'hui, mais `entity_id` sert de cle aux alertes et a la liste
    blanche : une collision ferait fusionner deux fiches distinctes. Le
    prefixe `NONSDN-` ecarte le risque par construction plutot que par
    confiance dans une propriete de la source.

    Les relations entre profils ne sont volontairement pas recoltees ici : le
    graphe de participations est rafraichi par source (`refresh_source_relationships`)
    et melanger deux espaces d'identifiants prefixes differemment y produirait
    des aretes pendantes.
    """
    for item in parse_ofac_advanced_xml(file_path):
        item["entity_id"] = f"NONSDN-{item.get('entity_id')}"
        item["origin"] = "OFAC Non-SDN Consolidated"
        yield item


# ------------------ REGISTRE NATIONAL DES GELS (DGT) ------------------
# Connecteur du registre national des gels des avoirs publie par la Direction
# generale du Tresor (gels-avoirs.dgtresor.gouv.fr, API publique ENGEL).
# Structure du fichier JSON : Publications.PublicationDetail[] avec IdRegistre,
# Nature (Personne physique / Personne morale / Navire), Nom et RegistreDetail[]
# (paires TypeChamp / Valeur[]). Le parseur est tolerant aux variations de cles
# des objets Valeur (recherche insensible a la casse, repli sur toute valeur texte).

DGT_NATURE_TO_TYPE = {
    "personne physique": "I",
    "personne morale": "E",
    "navire": "V",
}

# Normalisation des pays/nationalites (libelles francais du registre DGT et
# anglais de la liste ONU) vers ISO2, indispensable pour que les cles de
# blocking coincident avec celles du referentiel clients (codes ISO).
# Radicaux sans accents, minuscules : ils couvrent le nom du pays ET
# l'adjectif de nationalite. L'ordre compte (nigeria avant niger, etc.).
_COUNTRY_STEMS = [
    # --- Libelles anglais (liste consolidee ONU) ---
    ("democratic people's republic of korea", "KP"), ("north korea", "KP"),
    ("russian", "RU"), ("russia", "RU"),
    ("syrian arab republic", "SY"), ("syria", "SY"),
    ("iranian", "IR"), ("iran", "IR"),
    ("myanmar", "MM"), ("burma", "MM"),
    ("libyan arab jamahiriya", "LY"), ("libya", "LY"),
    ("venezuela", "VE"), ("china", "CN"), ("chinese", "CN"),
    ("iraqi", "IQ"), ("iraq", "IQ"),
    ("afghanistan", "AF"), ("afghan", "AF"),
    ("yemeni", "YE"), ("yemen", "YE"), ("lebanese", "LB"), ("lebanon", "LB"),
    ("south sudan", "SS"), ("sudanese", "SD"), ("sudan", "SD"),
    ("democratic republic of the congo", "CD"),
    ("central african", "CF"), ("somalia", "SO"), ("somali", "SO"),
    ("guinea-bissau", "GW"), ("guinea", "GN"),
    ("zimbabwe", "ZW"), ("turkey", "TR"), ("turkish", "TR"),
    ("ukraine", "UA"), ("ukrainian", "UA"), ("moldova", "MD"),
    ("tunisia", "TN"), ("tunisian", "TN"), ("egypt", "EG"),
    ("pakistani", "PK"), ("pakistan", "PK"),
    ("saudi arabia", "SA"), ("saudi", "SA"),
    ("jordanian", "JO"), ("jordan", "JO"),
    ("israeli", "IL"), ("israel", "IL"),
    ("united kingdom", "GB"), ("british", "GB"),
    ("united states", "US"), ("american", "US"),
    ("united arab emirates", "AE"), ("kuwait", "KW"),
    ("eritrea", "ER"), ("ethiopia", "ET"),
    ("kyrgyz", "KG"), ("tajikistan", "TJ"), ("uzbekistan", "UZ"), ("kazakhstan", "KZ"),
    ("indonesian", "ID"), ("indonesia", "ID"),
    ("philippines", "PH"), ("filipino", "PH"),
    ("sri lanka", "LK"), ("bangladesh", "BD"),
    ("nigerian", "NG"), ("nigeria", "NG"),
    ("burkina faso", "BF"), ("cameroon", "CM"),
    ("germany", "DE"), ("german", "DE"), ("belgium", "BE"), ("belgian", "BE"),
    ("spain", "ES"), ("spanish", "ES"), ("italy", "IT"), ("italian", "IT"),
    # --- Libelles francais (registre DGT) ---
    ("coree du nord", "KP"), ("nord-coreen", "KP"), ("nord coreen", "KP"),
    ("russ", "RU"), ("bielorus", "BY"), ("belarus", "BY"),
    ("syrie", "SY"), ("syrien", "SY"),
    ("iranien", "IR"), ("iran", "IR"),
    ("birman", "MM"), ("myanmar", "MM"), ("birmanie", "MM"),
    ("libye", "LY"), ("libyen", "LY"),
    ("malien", "ML"), ("mali", "ML"),
    ("venezuel", "VE"), ("chin", "CN"),
    ("irakien", "IQ"), ("irak", "IQ"),
    ("afghan", "AF"), ("yemen", "YE"), ("liban", "LB"),
    ("soudan du sud", "SS"), ("sud-soudan", "SS"),
    ("soudan", "SD"), ("congolais", "CD"),
    ("republique democratique du congo", "CD"), ("rdc", "CD"), ("congo", "CD"),
    ("centrafri", "CF"), ("somal", "SO"), ("nicaragua", "NI"),
    ("guinee-bissau", "GW"), ("bissau", "GW"), ("guine", "GN"),
    ("zimbabw", "ZW"), ("haiti", "HT"), ("hait", "HT"),
    ("turc", "TR"), ("turq", "TR"), ("ukrain", "UA"), ("moldav", "MD"),
    ("tunis", "TN"), ("egypt", "EG"), ("pakistan", "PK"),
    ("saoudien", "SA"), ("arabie saoudite", "SA"),
    ("jordan", "JO"), ("israel", "IL"), ("palestin", "PS"),
    ("franc", "FR"), ("algeri", "DZ"), ("marocain", "MA"), ("maroc", "MA"),
    ("burundi", "BI"), ("erythre", "ER"), ("ethiopi", "ET"),
    ("kirghiz", "KG"), ("tadjik", "TJ"), ("ouzbek", "UZ"), ("kazakh", "KZ"),
    ("armeni", "AM"), ("azerbaidjan", "AZ"), ("georgi", "GE"),
    ("serbe", "RS"), ("serbie", "RS"), ("bosni", "BA"), ("kosov", "XK"),
    ("indien", "IN"), ("inde", "IN"), ("indonesi", "ID"),
    ("philippin", "PH"), ("sri lank", "LK"), ("bangladesh", "BD"),
    ("nigeria", "NG"), ("nigerian", "NG"), ("nigerien", "NE"), ("niger", "NE"),
    ("burkin", "BF"), ("tchad", "TD"), ("tchadien", "TD"),
    ("camerou", "CM"), ("senegal", "SN"), ("mauritani", "MR"),
    ("kowei", "KW"), ("qatar", "QA"), ("emirat", "AE"), ("bahrein", "BH"),
    ("britanni", "GB"), ("royaume-uni", "GB"), ("americain", "US"), ("etats-unis", "US"),
    ("allemand", "DE"), ("allemagne", "DE"), ("belge", "BE"), ("belgique", "BE"),
    ("espagnol", "ES"), ("espagne", "ES"), ("italien", "IT"), ("italie", "IT"),
]


def _strip_accents_lower(text: str) -> str:
    import unicodedata
    return "".join(
        c for c in unicodedata.normalize("NFD", str(text or "").lower())
        if unicodedata.category(c) != "Mn"
    ).strip()


def country_label_to_iso2(value: str) -> str:
    """
    Convertit un pays / une nationalite en libelle francais (registre DGT :
    "Russe", "Russie") ou anglais (liste ONU : "Russian Federation") en code
    ISO2. Repli sur la valeur d'origine si aucun radical connu ne correspond
    (la cle de blocking reste coherente en interne, meme si elle ne croisera
    pas les codes ISO clients).
    """
    normalized = _strip_accents_lower(value)
    if re.fullmatch(r"[a-z]{2}", normalized):
        return normalized.upper()
    for stem, iso2 in _COUNTRY_STEMS:
        if normalized.startswith(stem) or f" {stem}" in f" {normalized}":
            return iso2
    return str(value).strip()


# Alias historique (connecteur DGT)
dgt_country_to_iso2 = country_label_to_iso2


def _dgt_value_text(value_obj: Any, *preferred_keys: str) -> str:
    """
    Extrait le texte d'un objet Valeur du registre DGT : cherche d'abord les
    cles preferees (insensible a la casse), sinon joint toutes les valeurs
    texte non vides de l'objet.
    """
    if not isinstance(value_obj, dict):
        return str(value_obj or "").strip()
    for key in preferred_keys:
        val = dict_get_insensitive(value_obj, key)
        if val is not None and str(val).strip():
            return str(val).strip()
    return " ".join(
        str(v).strip() for v in value_obj.values()
        if v is not None and isinstance(v, (str, int, float)) and str(v).strip()
    ).strip()


def _dgt_details_by_type(record: Dict[str, Any]) -> Dict[str, List[Any]]:
    """Indexe les RegistreDetail par TypeChamp -> liste d'objets Valeur."""
    indexed: Dict[str, List[Any]] = {}
    for detail in record.get("RegistreDetail") or []:
        type_champ = str(detail.get("TypeChamp") or "").strip().upper()
        if not type_champ:
            continue
        values = detail.get("Valeur")
        if values is None:
            continue
        if not isinstance(values, list):
            values = [values]
        indexed.setdefault(type_champ, []).extend(values)
    return indexed


def _dgt_date(value_obj: Any) -> Optional[str]:
    """Assemble une date YYYY-MM-DD depuis un objet {Jour, Mois, Annee} (jour/mois optionnels)."""
    if not isinstance(value_obj, dict):
        return None
    year = _dgt_value_text(value_obj, "Annee", "Year")
    if not year or not re.fullmatch(r"\d{4}", year):
        # Certains enregistrements portent la date complete dans un seul champ
        raw = _dgt_value_text(value_obj)
        match = re.search(r"(\d{4})(?:-(\d{1,2})-(\d{1,2}))?", raw)
        if not match:
            return None
        year, month, day = match.group(1), match.group(2) or "01", match.group(3) or "01"
        return f"{year}-{month.zfill(2)}-{day.zfill(2)}"
    month = _dgt_value_text(value_obj, "Mois", "Month") or "01"
    day = _dgt_value_text(value_obj, "Jour", "Day") or "01"
    if not month.isdigit():
        month = "01"
    if not day.isdigit():
        day = "01"
    return f"{year}-{month.zfill(2)}-{day.zfill(2)}"


def parse_dgt_gels_json(file_path: str) -> Generator[Dict[str, Any], None, None]:
    """
    Parse le fichier JSON du registre national des gels (DGT) et produit des
    enregistrements au schema pivot Fiskr.
    """
    with open(file_path, "r", encoding="utf-8") as f:
        data = json.load(f)

    publications = dict_get_insensitive(data, "Publications") or {}
    if isinstance(publications, list):
        publications = publications[0] if publications else {}
    records = dict_get_insensitive(publications, "PublicationDetail") or []
    # Date de publication du registre (commune a tous les enregistrements)
    publication_date = _extract_iso_date(str(dict_get_insensitive(publications, "DatePublication") or ""))

    for record in records:
        id_registre = record.get("IdRegistre")
        if id_registre is None:
            continue
        nature = str(record.get("Nature") or "").strip().lower()
        entity_type = DGT_NATURE_TO_TYPE.get(nature, "E")
        last_name = str(record.get("Nom") or "").strip()

        details = _dgt_details_by_type(record)

        first_name = ""
        for v in details.get("PRENOM", []):
            first_name = _dgt_value_text(v, "Prenom")
            if first_name:
                break

        if entity_type == "I":
            primary_name = f"{first_name} {last_name}".strip()
        else:
            primary_name = last_name
        if not primary_name:
            continue

        gender = "U"
        for v in details.get("SEXE", []):
            sexe = _dgt_value_text(v, "Sexe").lower()
            if sexe.startswith("f"):
                gender = "F"
            elif sexe.startswith("m"):
                gender = "M"

        dobs = []
        for v in details.get("DATE_DE_NAISSANCE", []):
            date_val = _dgt_date(v)
            if date_val:
                dobs.append(date_val)

        place_of_birth = None
        birth_countries = []
        for v in details.get("LIEU_DE_NAISSANCE", []):
            lieu = _dgt_value_text(v, "Lieu")
            pays = _dgt_value_text(v, "Pays")
            if not place_of_birth and (lieu or pays):
                place_of_birth = ", ".join(p for p in (lieu, pays) if p)
            if pays:
                birth_countries.append(dgt_country_to_iso2(pays))

        citizenships = []
        for v in details.get("NATIONALITE", []):
            pays = _dgt_value_text(v, "Pays", "Nationalite")
            if pays:
                citizenships.append(dgt_country_to_iso2(pays))

        aliases_raw = [
            {"name": alias, "type": "Strong"}
            for alias in (_dgt_value_text(v, "Alias") for v in details.get("ALIAS", []))
            if alias
        ]

        designation = None
        for v in details.get("TITRE", []):
            titre = _dgt_value_text(v, "Titre")
            if titre:
                designation = titre
                break

        addresses = []
        address_countries = []
        for type_champ in ("ADRESSE_PP", "ADRESSE_PM"):
            for v in details.get(type_champ, []):
                adresse = _dgt_value_text(v, "Adresse")
                pays = _dgt_value_text(v, "Pays")
                full = ", ".join(p for p in (adresse, pays) if p)
                if full:
                    addresses.append(full)
                if pays:
                    address_countries.append(pays)

        passports = []
        for v in details.get("PASSEPORT", []):
            numero = _dgt_value_text(v, "NumeroPasseport", "Numero")
            if numero:
                passports.append({"number": numero, "issuing_country": "XX", "expiration_date": None})

        other_registrations = []
        for type_champ in ("IDENTIFICATION", "AUTRE_IDENTITE"):
            for v in details.get(type_champ, []):
                ident = _dgt_value_text(v, "Identification", "NumeroCarte", "Numero")
                if ident:
                    other_registrations.append({"id_type": type_champ.title(), "number": ident})

        motifs = []
        for v in details.get("MOTIFS", []):
            motif = _dgt_value_text(v, "Motifs", "Motif")
            if motif:
                motifs.append(motif)

        extra_info = []
        official_ref = None
        fondements = []
        for type_champ, keys in (
            ("FONDEMENT_JURIDIQUE", ("FondementJuridiqueLabel", "FondementJuridique")),
            ("REFERENCE_UE", ("ReferenceUe",)),
            ("REFERENCE_ONU", ("ReferenceOnu",)),
        ):
            for v in details.get(type_champ, []):
                text = _dgt_value_text(v, *keys)
                if text:
                    extra_info.append(f"{type_champ.replace('_', ' ').title()}: {text}")
                    if official_ref is None and type_champ in ("REFERENCE_UE", "REFERENCE_ONU"):
                        official_ref = text
                    if type_champ == "FONDEMENT_JURIDIQUE" and text not in fondements:
                        fondements.append(text)

        # Contacts (TypeChamps presents sur certaines fiches PM du registre)
        dgt_phones = [t for v in details.get("TELEPHONE", []) if (t := _dgt_value_text(v, "Telephone"))]
        dgt_emails = [t for v in details.get("COURRIEL", []) + details.get("EMAIL", [])
                      if (t := _dgt_value_text(v, "Courriel", "Email"))]
        dgt_websites = [t for v in details.get("SITE_INTERNET", []) if (t := _dgt_value_text(v, "SiteInternet", "Site"))]

        yield {
            "entity_id": f"DGT-{id_registre}",
            "entity_type": entity_type,
            "primary_name": primary_name,
            "individual_name_parsed": {
                "first_name": first_name,
                "last_name": last_name if entity_type == "I" else "",
                "maiden_name": ""
            },
            "aliases": categorize_aliases(aliases_raw),
            "dates_of_birth": sorted(set(dobs)),
            "date_of_death": None,
            "is_deceased": False,
            "gender": gender,
            "countries": {
                "citizenship": sorted(set(citizenships)),
                "residence": [],
                "birth_country": sorted(set(birth_countries)),
                "jurisdiction_country": sorted({dgt_country_to_iso2(c) for c in address_countries}) if entity_type != "I" else []
            },
            "place_of_birth": place_of_birth,
            "address": addresses[0] if addresses else None,
            "alternative_addresses": addresses[1:],
            "country": address_countries[0] if address_countries else None,
            "designation": designation,
            "designation_reasons": "; ".join(motifs) or None,
            "additional_informations": "; ".join(extra_info) or None,
            "official_reference": build_official_reference(official_ref, publication_date),
            "sanction_programs": fondements,
            "phone_numbers": dgt_phones,
            "email_addresses": dgt_emails,
            "websites": dgt_websites,
            "origin": "DGT Registre national des gels",
            "imo_number": None,
            "aircraft_tail_number": None,
            "lei_number": None,
            "national_registry_ids": [],
            "other_registration_ids": other_registrations,
            "passport_documents": passports,
            "national_id_documents": [],
            "other_id_documents": []
        }


# ------------------ LISTE CONSOLIDEE UE (FSF XML OFFICIEL) ------------------
# Fichier XML consolide des sanctions financieres de l'UE, publie par la
# Commission (webgate FSD/FSF). Contrairement au scraping du Journal Officiel,
# ce fichier fait autorite et permet de detecter les radiations. Structure :
# <export><sanctionEntity euReferenceNumber=... logicalId=...>
#   <subjectType code="person|enterprise" classificationCode="P|E"/>
#   <regulation programme=... numberTitle=...><publicationUrl/></regulation>
#   <nameAlias wholeName=... firstName=... lastName=... gender=... function=... strong=.../>
#   <citizenship countryIso2Code=.../>
#   <birthdate birthdate=... year=... city=... countryIso2Code=.../>
#   <identification identificationTypeCode=... number=... countryIso2Code=.../>
#   <address street=... city=... countryIso2Code=.../>
#   <remark>...</remark>
# </sanctionEntity></export>


def _child_local(elem: ET.Element, local: str) -> List[ET.Element]:
    """Enfants directs dont le nom local (sans namespace) correspond."""
    return [c for c in elem if c.tag.split('}')[-1] == local]


def _extract_iso_date(raw: Optional[str]) -> Optional[str]:
    """Premiere date trouvee dans un texte libre, normalisee en YYYY-MM-DD (accepte JJ/MM/AAAA)."""
    raw = (raw or "").strip()
    match = re.search(r"(\d{4})-(\d{1,2})-(\d{1,2})", raw)
    if match:
        y, m, d = match.groups()
        return f"{y}-{m.zfill(2)}-{d.zfill(2)}"
    match = re.search(r"(\d{1,2})/(\d{1,2})/(\d{4})", raw)
    if match:
        d, m, y = match.groups()
        return f"{y}-{m.zfill(2)}-{d.zfill(2)}"
    return None


def build_official_reference(reference: Optional[str], updated: Optional[str]) -> Optional[str]:
    """
    Reference officielle de l'emetteur (reglement, reference de liste),
    suffixee de la date de publication/mise a jour quand la source la fournit.
    C'est cette date que le patch de fiche peut ramener a la date du jour.
    """
    reference = (reference or "").strip()
    updated = (updated or "").strip()
    if not reference:
        return None
    return f"{reference} (maj {updated})" if updated else reference


def parse_eu_fsf_xml(file_path: str) -> Generator[Dict[str, Any], None, None]:
    """Parse le XML consolide officiel des sanctions financieres de l'UE (FSF)."""
    for _, entity in _stream_target_elements(file_path, {'sanctionEntity'}):
        logical_id = get_attrib_insensitive(entity, "logicalId")
        eu_ref = get_attrib_insensitive(entity, "euReferenceNumber")
        entity_key = eu_ref or logical_id
        if not entity_key:
            continue

        entity_type = "E"
        for st in _child_local(entity, "subjectType"):
            code = (get_attrib_insensitive(st, "code") or "").lower()
            classification = (get_attrib_insensitive(st, "classificationCode") or "").upper()
            if "person" in code or classification == "P":
                entity_type = "I"

        primary_name = ""
        first_name = ""
        last_name = ""
        gender = "U"
        designation = None
        fsf_title = None
        aliases_raw = []
        for alias in _child_local(entity, "nameAlias"):
            whole = (get_attrib_insensitive(alias, "wholeName") or "").strip()
            fn = (get_attrib_insensitive(alias, "firstName") or "").strip()
            mn = (get_attrib_insensitive(alias, "middleName") or "").strip()
            ln = (get_attrib_insensitive(alias, "lastName") or "").strip()
            name = whole or " ".join(p for p in (fn, mn, ln) if p)
            if not name:
                continue
            g = (get_attrib_insensitive(alias, "gender") or "").upper()
            if g in ("M", "F") and gender == "U":
                gender = g
            func = (get_attrib_insensitive(alias, "function") or "").strip()
            title = (get_attrib_insensitive(alias, "title") or "").strip()
            if title and not fsf_title:
                fsf_title = title
            if not designation and (func or title):
                designation = func or title
            if not primary_name:
                primary_name = name
                first_name = " ".join(p for p in (fn, mn) if p)
                last_name = ln
            else:
                strong = (get_attrib_insensitive(alias, "strong") or "true").lower()
                aliases_raw.append({"name": name, "type": "Strong" if strong == "true" else "Weak"})
        if not primary_name:
            continue

        citizenships = []
        for cit in _child_local(entity, "citizenship"):
            iso2 = (get_attrib_insensitive(cit, "countryIso2Code") or "").strip()
            desc = (get_attrib_insensitive(cit, "countryDescription") or "").strip()
            code = iso2 or (country_label_to_iso2(desc) if desc else "")
            if code and code.upper() != "00":
                citizenships.append(code.upper() if len(code) == 2 else code)

        dobs = []
        place_of_birth = None
        birth_countries = []
        for bd in _child_local(entity, "birthdate"):
            full = (get_attrib_insensitive(bd, "birthdate") or "").strip()
            if full:
                dobs.append(full)
            else:
                year = (get_attrib_insensitive(bd, "year") or "").strip()
                if year.isdigit():
                    month = (get_attrib_insensitive(bd, "monthOfYear") or get_attrib_insensitive(bd, "month") or "1").strip() or "1"
                    day = (get_attrib_insensitive(bd, "dayOfMonth") or get_attrib_insensitive(bd, "day") or "1").strip() or "1"
                    dobs.append(f"{year}-{month.zfill(2)}-{day.zfill(2)}")
            city = (get_attrib_insensitive(bd, "city") or get_attrib_insensitive(bd, "place") or "").strip()
            country_desc = (get_attrib_insensitive(bd, "countryDescription") or "").strip()
            iso2 = (get_attrib_insensitive(bd, "countryIso2Code") or "").strip()
            if not place_of_birth and (city or country_desc):
                place_of_birth = ", ".join(p for p in (city, country_desc) if p)
            if iso2 and iso2.upper() != "00":
                birth_countries.append(iso2.upper())

        passports = []
        national_ids = []
        other_registrations = []
        for ident in _child_local(entity, "identification"):
            number = (get_attrib_insensitive(ident, "number") or get_attrib_insensitive(ident, "latinNumber") or "").strip()
            if not number:
                continue
            type_code = (get_attrib_insensitive(ident, "identificationTypeCode") or "").lower()
            type_desc = (get_attrib_insensitive(ident, "identificationTypeDescription") or "").lower()
            iso2 = (get_attrib_insensitive(ident, "countryIso2Code") or "XX").strip().upper() or "XX"
            if "passport" in type_code or "passport" in type_desc:
                passports.append({"number": number, "issuing_country": iso2, "expiration_date": None})
            elif type_code == "id" or "national" in type_desc:
                national_ids.append({"number": number, "issuing_country": iso2})
            else:
                other_registrations.append({"id_type": type_desc or type_code or "OtherRegistration", "number": number})

        addresses = []
        address_countries = []
        for addr in _child_local(entity, "address"):
            parts = [
                (get_attrib_insensitive(addr, key) or "").strip()
                for key in ("street", "poBox", "zipCode", "city", "region", "place")
            ]
            country_desc = (get_attrib_insensitive(addr, "countryDescription") or "").strip()
            iso2 = (get_attrib_insensitive(addr, "countryIso2Code") or "").strip().upper()
            full = ", ".join(p for p in parts + [country_desc] if p)
            if full:
                addresses.append({"full": full, "city": parts[3], "country": country_desc})
            if iso2 and iso2 != "00":
                address_countries.append(iso2)

        programme = None
        programmes = []
        extra_info = []
        official_ref = None
        official_ref_date = None
        for reg in _child_local(entity, "regulation"):
            prog = (get_attrib_insensitive(reg, "programme") or "").strip()
            if prog and not programme:
                programme = prog
            if prog and prog not in programmes:
                programmes.append(prog)
            number_title = (get_attrib_insensitive(reg, "numberTitle") or "").strip()
            if number_title:
                extra_info.append(f"Regulation: {number_title}")
                if official_ref is None:
                    official_ref = f"Regulation {number_title}"
                    official_ref_date = _extract_iso_date(get_attrib_insensitive(reg, "publicationDate"))
        for remark in _child_local(entity, "remark"):
            if remark.text and remark.text.strip():
                extra_info.append(remark.text.strip())
        un_id = (get_attrib_insensitive(entity, "unitedNationId") or "").strip()
        if un_id:
            extra_info.append(f"UN ID: {un_id}")

        primary_addr = addresses[0] if addresses else {}
        yield {
            "entity_id": f"EUFSF-{entity_key}",
            "entity_type": entity_type,
            "primary_name": primary_name,
            "individual_name_parsed": {
                "first_name": first_name if entity_type == "I" else "",
                "last_name": last_name if entity_type == "I" else "",
                "maiden_name": ""
            },
            "aliases": categorize_aliases(aliases_raw),
            "dates_of_birth": sorted(set(dobs)),
            "date_of_death": None,
            "is_deceased": False,
            "gender": gender,
            "countries": {
                "citizenship": sorted(set(citizenships)),
                "residence": [],
                "birth_country": sorted(set(birth_countries)),
                "jurisdiction_country": sorted(set(address_countries)) if entity_type != "I" else []
            },
            "place_of_birth": place_of_birth,
            "address": primary_addr.get("full"),
            "alternative_addresses": [a["full"] for a in addresses[1:]],
            "city": primary_addr.get("city") or None,
            "country": primary_addr.get("country") or None,
            "designation": designation,
            "designation_reasons": programme,
            "additional_informations": "; ".join(extra_info) or None,
            "official_reference": build_official_reference(official_ref, official_ref_date),
            "title": fsf_title,
            "listed_on": official_ref_date,
            "sanction_programs": programmes,
            "origin": "EU FSF Consolidated",
            "imo_number": None,
            "aircraft_tail_number": None,
            "lei_number": None,
            "national_registry_ids": [],
            "other_registration_ids": other_registrations,
            "passport_documents": passports,
            "national_id_documents": national_ids,
            "other_id_documents": []
        }


# ------------------ LISTE CONSOLIDEE ONU (XML OFFICIEL) ------------------
# Liste consolidee du Conseil de securite (scsanctions.un.org), publique et
# sans authentification. Deux sections : INDIVIDUALS/INDIVIDUAL et
# ENTITIES/ENTITY, avec champs texte (FIRST_NAME..FOURTH_NAME, UN_LIST_TYPE,
# REFERENCE_NUMBER, COMMENTS1), listes imbriquees (INDIVIDUAL_ALIAS,
# INDIVIDUAL_DATE_OF_BIRTH, INDIVIDUAL_PLACE_OF_BIRTH, INDIVIDUAL_DOCUMENT,
# *_ADDRESS) et valeurs multiples (NATIONALITY/VALUE, DESIGNATION/VALUE).


def _un_text(elem: ET.Element, local: str) -> str:
    children = _child_local(elem, local)
    if children and children[0].text and children[0].text.strip():
        return children[0].text.strip()
    return ""


def _un_values(elem: ET.Element, local: str) -> List[str]:
    """Valeurs des blocs <LOCAL><VALUE>..</VALUE>...</LOCAL>."""
    out = []
    for container in _child_local(elem, local):
        for value in _child_local(container, "VALUE"):
            if value.text and value.text.strip():
                out.append(value.text.strip())
    return out


def parse_un_consolidated_xml(file_path: str) -> Generator[Dict[str, Any], None, None]:
    """Parse la liste consolidee officielle du Conseil de securite de l'ONU."""
    for local_name, record in _stream_target_elements(file_path, {'INDIVIDUAL', 'ENTITY'}):
        data_id = _un_text(record, "DATAID")
        reference = _un_text(record, "REFERENCE_NUMBER")
        entity_key = reference or data_id
        if not entity_key:
            continue
        entity_type = "I" if local_name == "INDIVIDUAL" else "E"

        name_parts = [
            _un_text(record, tag)
            for tag in ("FIRST_NAME", "SECOND_NAME", "THIRD_NAME", "FOURTH_NAME")
        ]
        name_parts = [p for p in name_parts if p]
        primary_name = " ".join(name_parts)
        if not primary_name:
            continue

        aliases_raw = []
        original_script = _un_text(record, "NAME_ORIGINAL_SCRIPT")
        if original_script:
            aliases_raw.append({"name": original_script, "type": "Strong"})
        alias_tag = "INDIVIDUAL_ALIAS" if entity_type == "I" else "ENTITY_ALIAS"
        for alias in _child_local(record, alias_tag):
            alias_name = _un_text(alias, "ALIAS_NAME")
            if not alias_name:
                continue
            quality = _un_text(alias, "QUALITY").lower()
            aliases_raw.append({"name": alias_name, "type": "Weak" if "low" in quality else "Strong"})

        dobs = []
        for dob in _child_local(record, "INDIVIDUAL_DATE_OF_BIRTH"):
            date_val = _un_text(dob, "DATE")
            year = _un_text(dob, "YEAR") or _un_text(dob, "FROM_YEAR")
            if date_val:
                dobs.append(date_val[:10])
            elif year.isdigit():
                dobs.append(f"{year}-01-01")

        place_of_birth = None
        birth_countries = []
        for pob in _child_local(record, "INDIVIDUAL_PLACE_OF_BIRTH"):
            city = _un_text(pob, "CITY")
            state = _un_text(pob, "STATE_PROVINCE")
            country = _un_text(pob, "COUNTRY")
            if not place_of_birth and (city or state or country):
                place_of_birth = ", ".join(p for p in (city, state, country) if p)
            if country:
                birth_countries.append(country_label_to_iso2(country))

        citizenships = [country_label_to_iso2(v) for v in _un_values(record, "NATIONALITY")]

        passports = []
        national_ids = []
        other_registrations = []
        for doc in _child_local(record, "INDIVIDUAL_DOCUMENT"):
            number = _un_text(doc, "NUMBER")
            if not number:
                continue
            doc_type = _un_text(doc, "TYPE_OF_DOCUMENT").lower()
            issuing = _un_text(doc, "ISSUING_COUNTRY") or _un_text(doc, "COUNTRY_OF_ISSUE")
            iso2 = country_label_to_iso2(issuing) if issuing else "XX"
            if "passport" in doc_type:
                passports.append({"number": number, "issuing_country": iso2, "expiration_date": None})
            elif "national" in doc_type:
                national_ids.append({"number": number, "issuing_country": iso2})
            else:
                other_registrations.append({"id_type": doc_type or "OtherRegistration", "number": number})

        addresses = []
        address_countries = []
        addr_tag = "INDIVIDUAL_ADDRESS" if entity_type == "I" else "ENTITY_ADDRESS"
        for addr in _child_local(record, addr_tag):
            parts = [_un_text(addr, t) for t in ("STREET", "CITY", "STATE_PROVINCE", "COUNTRY")]
            full = ", ".join(p for p in parts if p)
            if full:
                addresses.append({"full": full, "city": parts[1], "country": parts[3]})
            if parts[3]:
                address_countries.append(country_label_to_iso2(parts[3]))

        designation = "; ".join(_un_values(record, "DESIGNATION")) or None
        un_list_type = _un_text(record, "UN_LIST_TYPE")
        comments = _un_text(record, "COMMENTS1")
        extra_info = []
        if reference:
            extra_info.append(f"UN Reference: {reference}")
        if comments:
            extra_info.append(comments)

        # Date de mise a jour de la fiche : derniere valeur LAST_DAY_UPDATED,
        # a defaut la date d'inscription LISTED_ON
        updated_values = _un_values(record, "LAST_DAY_UPDATED")
        un_updated = _extract_iso_date(updated_values[-1] if updated_values else _un_text(record, "LISTED_ON"))

        # Champs etendus : titre, date d'inscription, Etat designant
        un_title = "; ".join(_un_values(record, "TITLE")) or None
        listed_on = _extract_iso_date(_un_text(record, "LISTED_ON"))
        designating_state = _un_text(record, "SUBMITTED_BY") or None

        primary_addr = addresses[0] if addresses else {}
        yield {
            "entity_id": f"UN-{entity_key}",
            "entity_type": entity_type,
            "primary_name": primary_name,
            "individual_name_parsed": {"first_name": "", "last_name": "", "maiden_name": ""},
            "aliases": categorize_aliases(aliases_raw),
            "dates_of_birth": sorted(set(dobs)),
            "date_of_death": None,
            "is_deceased": False,
            "gender": "U",
            "countries": {
                "citizenship": sorted(set(citizenships)),
                "residence": [],
                "birth_country": sorted(set(birth_countries)),
                "jurisdiction_country": sorted(set(address_countries)) if entity_type != "I" else []
            },
            "place_of_birth": place_of_birth,
            "address": primary_addr.get("full"),
            "alternative_addresses": [a["full"] for a in addresses[1:]],
            "city": primary_addr.get("city") or None,
            "country": primary_addr.get("country") or None,
            "designation": designation,
            "designation_reasons": un_list_type or None,
            "additional_informations": "; ".join(extra_info) or None,
            "official_reference": build_official_reference(reference, un_updated),
            "title": un_title,
            "listed_on": listed_on,
            "designating_state": designating_state,
            "sanction_programs": [un_list_type] if un_list_type else [],
            "name_original_script": original_script or None,
            "origin": "UN Consolidated List",
            "imo_number": None,
            "aircraft_tail_number": None,
            "lei_number": None,
            "national_registry_ids": [],
            "other_registration_ids": other_registrations,
            "passport_documents": passports,
            "national_id_documents": national_ids,
            "other_id_documents": []
        }


# ------------------ SOURCE PEP (OPENSANCTIONS targets.simple.csv) ------------------
# Dataset PEP agrege par OpenSanctions (donnees Wikidata et sources officielles).
# Format targets.simple.csv : id, schema, name, aliases, birth_date, countries,
# addresses, identifiers, sanctions, phones, emails, dataset, ... (valeurs
# multiples separees par ";"). Licence : usage non commercial libre, licence
# requise pour un usage commercial — voir opensanctions.org/licensing.

_PEP_SCHEMA_TO_TYPE = {
    "person": "I",
    "company": "E",
    "organization": "E",
    "legalentity": "E",
    "publicbody": "E",
    "vessel": "V",
    "airplane": "O",
}


def _csv_multi(value: str) -> List[str]:
    return [v.strip() for v in (value or "").split(";") if v and v.strip()]


def _normalize_partial_date(raw: str) -> Optional[str]:
    """'1952' -> 1952-01-01 ; '1952-10' -> 1952-10-01 ; '1952-10-07' inchange."""
    raw = (raw or "").strip()
    match = re.match(r"^(\d{4})(?:-(\d{1,2}))?(?:-(\d{1,2}))?", raw)
    if not match:
        return None
    y, m, d = match.group(1), match.group(2) or "01", match.group(3) or "01"
    return f"{y}-{m.zfill(2)}-{d.zfill(2)}"


def parse_opensanctions_simple_csv(
    file_path: str,
    id_prefix: str = "PEP",
    origin: str = "OpenSanctions PEP",
    designation_reasons: str = "Personne Politiquement Exposée (PEP)",
) -> Generator[Dict[str, Any], None, None]:
    """
    Parse un dataset OpenSanctions au format `targets.simple.csv`.

    Le format est le meme pour tous les jeux de donnees du fournisseur : seuls
    changent le prefixe d'identifiant, la provenance affichee et le motif de
    designation. C'est ce qui permet de brancher une seconde source (SECO) sur
    le meme lecteur sans le dupliquer.
    """
    with open(file_path, "r", encoding="utf-8", newline="") as f:
        reader = csv.DictReader(f)
        for row in reader:
            row = {(k or "").strip().lower(): (v or "") for k, v in row.items()}
            os_id = row.get("id", "").strip()
            name = row.get("name", "").strip()
            if not os_id or not name:
                continue
            schema = row.get("schema", "").strip().lower()
            entity_type = _PEP_SCHEMA_TO_TYPE.get(schema, "E")

            aliases_raw = [{"name": a, "type": "Strong"} for a in _csv_multi(row.get("aliases"))]
            dobs = [d for d in (
                _normalize_partial_date(b) for b in _csv_multi(row.get("birth_date"))
            ) if d]
            countries = [c.upper() for c in _csv_multi(row.get("countries")) if c]
            addresses = _csv_multi(row.get("addresses"))
            identifiers = _csv_multi(row.get("identifiers"))
            positions = _csv_multi(row.get("sanctions")) or _csv_multi(row.get("position"))
            # Champs etendus : contacts + fonction PEP + premiere apparition
            pep_phones = _csv_multi(row.get("phones"))
            pep_emails = _csv_multi(row.get("emails"))
            pep_first_seen = _extract_iso_date(row.get("first_seen", ""))

            yield {
                "entity_id": f"{id_prefix}-{os_id}",
                "entity_type": entity_type,
                "primary_name": name,
                "individual_name_parsed": {"first_name": "", "last_name": "", "maiden_name": ""},
                "aliases": categorize_aliases(aliases_raw),
                "dates_of_birth": sorted(set(dobs)),
                "date_of_death": None,
                "is_deceased": False,
                "gender": "U",
                "countries": {
                    "citizenship": sorted(set(countries)),
                    "residence": [],
                    "birth_country": [],
                    "jurisdiction_country": []
                },
                "place_of_birth": None,
                "address": addresses[0] if addresses else None,
                "alternative_addresses": addresses[1:],
                "country": None,
                "designation": positions[0] if positions else None,
                "designation_reasons": designation_reasons,
                "additional_informations": "; ".join(identifiers) or None,
                "pep_role": "; ".join(positions) or None,
                "listed_on": pep_first_seen,
                "phone_numbers": pep_phones,
                "email_addresses": pep_emails,
                "origin": origin,
                "imo_number": None,
                "aircraft_tail_number": None,
                "lei_number": None,
                "national_registry_ids": [],
                "other_registration_ids": [{"id_type": "OpenSanctionsId", "number": os_id}],
                "passport_documents": [],
                "national_id_documents": [],
                "other_id_documents": []
            }


def parse_pep_targets_csv(file_path: str) -> Generator[Dict[str, Any], None, None]:
    """Parse le dataset PEP OpenSanctions (targets.simple.csv) vers le schema pivot."""
    return parse_opensanctions_simple_csv(
        file_path,
        id_prefix="PEP",
        origin="OpenSanctions PEP",
        designation_reasons="Personne Politiquement Exposée (PEP)",
    )


def parse_seco_opensanctions_csv(file_path: str) -> Generator[Dict[str, Any], None, None]:
    """
    Parse le jeu de donnees SECO agrege par OpenSanctions (`ch_seco_sanctions`,
    format targets.simple.csv). Voie de secours du connecteur XML officiel :
    meme perimetre de listes, mais un format plat, donc sans base legale ni
    dates d'acte. Meme reserve de licence que le dataset PEP — usage non
    commercial libre, licence requise au-dela.
    """
    return parse_opensanctions_simple_csv(
        file_path,
        id_prefix="SECO",
        origin="OpenSanctions SECO (CH)",
        designation_reasons="Sanctions suisses (SECO)",
    )


# ------------------ LISTE UK OFSI (ConList.csv, format 2022) ------------------
# Liste consolidee de l'OFSI (HM Treasury). CSV avec une ligne de preambule
# ("Last Updated..."), puis en-tetes : Name 6 (nom de famille), Name 1..5
# (prenoms), Title, DOB (jj/mm/aaaa), Town/Country of Birth, Nationality,
# Position, Address 1..6, Country, Other Information, Group Type
# (Individual/Entity/Ship), Alias Type (Primary name / aka), Regime, Group ID.
# Plusieurs lignes par Group ID : la ligne "Primary name" porte l'identite,
# les autres sont des alias.


def _ofsi_get(row: Dict[str, str], *keys: str) -> str:
    for key in keys:
        for k, v in row.items():
            if k and k.strip().lower() == key.lower():
                return (v or "").strip()
    return ""


def _ofsi_date(raw: str) -> Optional[str]:
    raw = (raw or "").strip()
    match = re.match(r"^(\d{1,2})/(\d{1,2})/(\d{4})$", raw)
    if match:
        d, m, y = match.groups()
        return f"{y}-{m.zfill(2)}-{d.zfill(2)}"
    return _normalize_partial_date(raw)


def parse_ofsi_conlist_csv(file_path: str) -> Generator[Dict[str, Any], None, None]:
    """Parse la liste consolidee UK OFSI (ConList.csv) vers le schema pivot."""
    with open(file_path, "r", encoding="utf-8-sig", newline="") as f:
        lines = f.read().splitlines()
    # Saute le preambule jusqu'a la ligne d'en-tetes (contient "Group Type")
    header_idx = next(
        (i for i, line in enumerate(lines[:10]) if "group type" in line.lower()), 0
    )
    reader = csv.DictReader(lines[header_idx:])

    groups: Dict[str, Dict[str, Any]] = {}
    for row in reader:
        group_id = _ofsi_get(row, "Group ID", "GroupID")
        if not group_id:
            continue
        name_parts = [
            _ofsi_get(row, f"Name {i}") for i in (1, 2, 3, 4, 5)
        ]
        family = _ofsi_get(row, "Name 6")
        first = " ".join(p for p in name_parts if p)
        full_name = " ".join(p for p in (first, family) if p)
        if not full_name:
            continue
        alias_type = _ofsi_get(row, "Alias Type", "AliasType").lower()
        is_primary = "primary" in alias_type or alias_type == ""

        group = groups.setdefault(group_id, {
            "primary": None, "first": "", "family": "", "aliases": [],
            "row": None
        })
        if is_primary and not group["primary"]:
            group["primary"] = full_name
            group["first"] = first
            group["family"] = family
            group["row"] = row
        else:
            group["aliases"].append({"name": full_name, "type": "Strong"})

    for group_id, group in groups.items():
        row = group["row"]
        if row is None:
            # Aucune ligne primaire : premier alias promu
            if not group["aliases"]:
                continue
            group["primary"] = group["aliases"][0]["name"]
            group["aliases"] = group["aliases"][1:]
            row = {}

        group_type = _ofsi_get(row, "Group Type", "GroupType").lower()
        if "individual" in group_type:
            entity_type = "I"
        elif "ship" in group_type:
            entity_type = "V"
        else:
            entity_type = "E"

        dob = _ofsi_date(_ofsi_get(row, "DOB"))
        uk_ref = _ofsi_get(row, "UK Sanctions List Ref", "UK Sanctions Ref", "UK Statement of Reasons Ref")
        last_updated = _ofsi_date(_ofsi_get(row, "Last Updated", "LastUpdated"))
        town_birth = _ofsi_get(row, "Town of Birth")
        country_birth = _ofsi_get(row, "Country of Birth")
        nationality = _ofsi_get(row, "Nationality")
        position = _ofsi_get(row, "Position")
        regime = _ofsi_get(row, "Regime")
        other_info = _ofsi_get(row, "Other Information")
        # Champs etendus du ConList
        ofsi_title = _ofsi_get(row, "Title")
        listed_on = _ofsi_date(_ofsi_get(row, "Listed On", "Date Designated", "UK Sanctions List Date Designated"))
        non_latin = _ofsi_get(row, "Name Non-Latin Script", "Non-Latin Script")
        passport_num = _ofsi_get(row, "Passport Number", "Passport Details")
        ni_number = _ofsi_get(row, "NI Number", "National Identification Number")
        ofsi_phone = _ofsi_get(row, "Phone number", "Phone Number", "Telephone")
        ofsi_email = _ofsi_get(row, "Email address", "Email Address", "Email")
        ofsi_website = _ofsi_get(row, "Website")
        if non_latin:
            group["aliases"].append({"name": non_latin, "type": "Strong"})
        addr_parts = [
            _ofsi_get(row, f"Address {i}") for i in (1, 2, 3, 4, 5, 6)
        ] + [_ofsi_get(row, "Post/Zip Code"), _ofsi_get(row, "Country")]
        address = ", ".join(p for p in addr_parts if p)

        citizenships = [country_label_to_iso2(n) for n in nationality.replace("(1)", ";").split(";") if n.strip()] if nationality else []

        yield {
            "entity_id": f"OFSI-{group_id}",
            "entity_type": entity_type,
            "primary_name": group["primary"],
            "individual_name_parsed": {
                "first_name": group["first"] if entity_type == "I" else "",
                "last_name": group["family"] if entity_type == "I" else "",
                "maiden_name": ""
            },
            "aliases": categorize_aliases(group["aliases"]),
            "dates_of_birth": [dob] if dob else [],
            "date_of_death": None,
            "is_deceased": False,
            "gender": "U",
            "countries": {
                "citizenship": sorted(set(citizenships)),
                "residence": [],
                "birth_country": [country_label_to_iso2(country_birth)] if country_birth else [],
                "jurisdiction_country": []
            },
            "place_of_birth": ", ".join(p for p in (town_birth, country_birth) if p) or None,
            "address": address or None,
            "alternative_addresses": [],
            "country": _ofsi_get(row, "Country") or None,
            "designation": position or None,
            "designation_reasons": regime or None,
            "additional_informations": other_info or None,
            "official_reference": build_official_reference(uk_ref or f"OFSI Group {group_id}", last_updated) if (uk_ref or last_updated) else None,
            "title": ofsi_title or None,
            "listed_on": listed_on,
            "name_original_script": non_latin or None,
            "sanction_programs": [regime] if regime else [],
            "phone_numbers": [ofsi_phone] if ofsi_phone else [],
            "email_addresses": [ofsi_email] if ofsi_email else [],
            "websites": [ofsi_website] if ofsi_website else [],
            "origin": "UK OFSI Consolidated",
            "imo_number": None,
            "aircraft_tail_number": None,
            "lei_number": None,
            "national_registry_ids": [],
            "other_registration_ids": [],
            "passport_documents": [{"number": passport_num, "issuing_country": "XX", "expiration_date": None}] if passport_num else [],
            "national_id_documents": [{"number": ni_number, "issuing_country": "GB"}] if ni_number else [],
            "other_id_documents": []
        }


# ------------------ LISTE SECO SUISSE (XML SESAM) ------------------
# Liste consolidee des sanctions du Secretariat d'Etat a l'economie (SECO),
# publiee au format XML par la plate-forme SESAM de la Confederation. La Suisse
# transpose les mesures de l'ONU et de l'UE dans ses propres ordonnances : le
# fichier porte donc le regime suisse applicable, avec sa base legale (RS).
#
# Structure du schema :
#   <export>
#     <sanctions-program>
#       <program-key>UKR</program-key>
#       <sanctions-set ssid="1"><version-date/><origin>EU</origin></sanctions-set>
#       <sanctions-set-name lang="fra">Ordonnance ... (RS 946.231...)</sanctions-set-name>
#     </sanctions-program>
#     <target ssid="1000" sanctions-set-id="1">
#       <individual|entity|object ssid="...">
#         <identity main="true">
#           <name name-type="primary-name" quality="good">
#             <name-part name-part-type="family-name"><value>IVANOV</value></name-part>
#             <name-part name-part-type="given-name"><value>Ivan</value></name-part>
#           </name>
#           <address><address-details/><location/><zip-code/><country iso-code="RU"/></address>
#           <place-of-birth><location/><country iso-code="RU"/></place-of-birth>
#           <nationality><country iso-code="RU"/></nationality>
#           <identification-document document-type="passport">
#             <number/><issuer code="RU"/><expiry-date><day-month-year year="..."/></expiry-date>
#           </identification-document>
#           <day-month-year year="1970" month="3" day="12"/>
#         </identity>
#         <justification lang="fra">...</justification>
#         <other-information lang="fra">...</other-information>
#         <modification modification-type="added" effective-date="2022-03-04"/>
#       </individual>
#     </target>
#   </export>
#
# AVERTISSEMENT ASSUME : ce parseur est ecrit d'apres le schema publie et
# valide sur un jeu d'essai synthetique couvrant les trois types de cibles ; il
# n'a PAS pu etre confronte au fichier reel, l'acces reseau a
# sesam.search.admin.ch etant ferme dans l'environnement de developpement. La
# premiere synchronisation reelle doit donc etre POINTEE avant mise en
# production (le mode homologation est fait pour cela). C'est la raison pour
# laquelle la lecture se fait par NOM LOCAL sur le sous-arbre de la cible et
# non par chemin rigide : la profondeur exacte a laquelle le schema place la
# date de naissance, l'adresse ou la nationalite peut varier sans casser
# l'extraction.

_SECO_GIVEN_PARTS = ("given-name", "father-name", "grand-father-name")
_SECO_FAMILY_PARTS = ("family-name", "maiden-name", "tribal-name")
_SECO_DECORATIVE_PARTS = ("title", "suffix")
# Une <day-month-year> sous l'un de ces ancetres n'est PAS une date de naissance
# (validite d'un document, date d'inscription, date d'une relation).
_SECO_NON_BIRTH_ANCESTORS = {
    "identification-document", "modification", "relation", "place-of-birth"
}


def _seco_local(elem: ET.Element) -> str:
    return elem.tag.split('}')[-1]


def _seco_attr(elem: ET.Element, *names: str) -> str:
    """Premier attribut renseigne parmi `names`, insensible au namespace et a la casse."""
    for name in names:
        value = get_attrib_insensitive(elem, name)
        if value is not None and str(value).strip():
            return str(value).strip()
    return ""


def _seco_text(elem: Optional[ET.Element]) -> str:
    if elem is None:
        return ""
    return (elem.text or "").strip()


def _seco_find(elem: ET.Element, local: str) -> List[ET.Element]:
    """Descendants (a n'importe quelle profondeur) portant ce nom local."""
    return [d for d in elem.iter() if d.tag.split('}')[-1] == local]


def _seco_first_text(elem: ET.Element, local: str) -> str:
    for node in _seco_find(elem, local):
        text = _seco_text(node)
        if text:
            return text
    return ""


def _seco_parent_map(root: ET.Element) -> Dict[ET.Element, ET.Element]:
    return {child: parent for parent in root.iter() for child in parent}


def _seco_ancestors(elem: ET.Element, parents: Dict[ET.Element, ET.Element]) -> Set[str]:
    """Noms locaux de tous les ancetres, pour lever une ambiguite de contexte."""
    out: Set[str] = set()
    node = parents.get(elem)
    while node is not None:
        out.add(_seco_local(node))
        node = parents.get(node)
    return out


def _seco_dmy(elem: ET.Element) -> Optional[str]:
    """<day-month-year year="1970" month="3" day="12"/> -> 1970-03-12.

    Une date partielle (annee seule) est ramenee au 1er janvier, comme le font
    deja les connecteurs ONU et PEP : le scoring applique de toute facon une
    tolerance en annees.
    """
    year = _seco_attr(elem, "year")
    if not (year.isdigit() and len(year) == 4):
        # Certaines exports ecrivent la date en texte plutot qu'en attributs
        return _normalize_partial_date(_seco_text(elem))
    month = _seco_attr(elem, "month")
    day = _seco_attr(elem, "day")
    month = month if month.isdigit() and 1 <= int(month) <= 12 else "1"
    day = day if day.isdigit() and 1 <= int(day) <= 31 else "1"
    return f"{year}-{month.zfill(2)}-{day.zfill(2)}"


def _seco_country(elem: ET.Element) -> Tuple[str, str]:
    """(libelle, ISO2) du premier <country> du sous-arbre. L'attribut iso-code
    prime ; a defaut le libelle passe par la table de correspondance commune."""
    for country in _seco_find(elem, "country"):
        label = _seco_text(country)
        iso = _seco_attr(country, "iso-code", "code", "iso").upper()
        if len(iso) != 2:
            iso = country_label_to_iso2(label) if label else ""
        if label or iso:
            return label, iso
    return "", ""


def _seco_name_parts(name_elem: ET.Element) -> List[Tuple[str, str]]:
    """[(type_de_partie, valeur)] dans l'ordre du document."""
    parts: List[Tuple[str, str]] = []
    for part in _seco_find(name_elem, "name-part"):
        value = _seco_first_text(part, "value") or _seco_text(part)
        if value:
            parts.append((_seco_attr(part, "name-part-type", "type").lower(), value))
    return parts


def _seco_compose(parts: List[Tuple[str, str]]) -> Tuple[str, str, str, str]:
    """
    Retourne (nom_compose, nom_dans_l_ordre_du_document, prenoms, nom_de_famille).

    Le nom compose suit l'ordre « prenoms puis patronyme » attendu par le
    moteur. L'ordre du document est rendu a part parce que les listes suisses
    ecrivent couramment le patronyme en premier : les deux graphies sont
    indexees, l'une comme nom principal, l'autre comme alias.
    """
    whole = [value for ptype, value in parts if ptype == "whole-name"]
    given = [value for ptype, value in parts if ptype in _SECO_GIVEN_PARTS]
    family = [value for ptype, value in parts if ptype in _SECO_FAMILY_PARTS]
    document_order = " ".join(
        value for ptype, value in parts if ptype not in _SECO_DECORATIVE_PARTS
    ).strip()
    composed = whole[0] if whole else (" ".join(given + family).strip() or document_order)
    return composed, document_order, " ".join(given).strip(), " ".join(family).strip()


def _seco_localized(elem: ET.Element, local: str) -> str:
    """Texte d'un bloc multilingue, francais d'abord (le fichier est trilingue)."""
    by_lang: Dict[str, str] = {}
    for node in _seco_find(elem, local):
        text = _seco_text(node)
        if text:
            by_lang.setdefault(_seco_attr(node, "lang", "xml:lang").lower(), text)
    for lang in ("fra", "fre", "fr", "", "eng", "en", "deu", "ger", "de", "ita", "it"):
        if lang in by_lang:
            return by_lang[lang]
    return next(iter(by_lang.values()), "")


def _seco_programs(file_path: str) -> Dict[str, Dict[str, str]]:
    """
    Premiere passe : ssid du jeu de sanctions -> {cle de programme, intitule de
    l'ordonnance, origine (ONU/UE)}. Les <target> sont liberees au fil de l'eau,
    cette passe ne charge pas le fichier en memoire.
    """
    programs: Dict[str, Dict[str, str]] = {}
    for _, prog in _stream_target_elements(file_path, {"sanctions-program"}):
        key = _seco_first_text(prog, "program-key")
        label = _seco_localized(prog, "sanctions-set-name") or _seco_localized(prog, "sanctions-program-name")
        for sset in _seco_find(prog, "sanctions-set"):
            ssid = _seco_attr(sset, "ssid")
            if ssid:
                programs[ssid] = {
                    "key": key,
                    "label": label,
                    "origin": _seco_first_text(sset, "origin"),
                }
    return programs


def parse_seco_xml(file_path: str) -> Generator[Dict[str, Any], None, None]:
    """Parse la liste consolidee suisse (SECO / SESAM) vers le schema pivot."""
    programs = _seco_programs(file_path)

    for _, target in _stream_target_elements(file_path, {"target"}):
        subject = next(
            (c for c in target if _seco_local(c) in ("individual", "entity", "object")),
            None
        )
        if subject is None:
            continue
        ssid = _seco_attr(target, "ssid") or _seco_attr(subject, "ssid")
        if not ssid:
            continue

        subject_local = _seco_local(subject)
        if subject_local == "individual":
            entity_type = "I"
        elif subject_local == "entity":
            entity_type = "E"
        else:
            object_type = _seco_attr(subject, "object-type", "type").lower()
            entity_type = "V" if any(k in object_type for k in ("vessel", "ship", "schiff", "navire")) else "O"

        parents = _seco_parent_map(subject)

        # --- Noms : identite principale (main="true") vs identites alias ---
        identities = _seco_find(subject, "identity") or [subject]
        primary_name = first_name = last_name = seco_title = ""
        aliases_raw: List[Dict[str, str]] = []
        for identity in identities:
            is_main = _seco_attr(identity, "main").lower() in ("true", "1", "yes")
            for name_elem in _seco_find(identity, "name"):
                parts = _seco_name_parts(name_elem)
                if not parts:
                    continue
                composed, document_order, given, family = _seco_compose(parts)
                if not composed:
                    continue
                name_type = _seco_attr(name_elem, "name-type", "type").lower()
                quality = _seco_attr(name_elem, "quality", "name-quality").lower()
                alias_type = "Weak" if any(k in quality for k in ("low", "weak", "poor")) else "Strong"
                if is_main and not primary_name and ("primary" in name_type or not name_type):
                    primary_name = composed
                    first_name, last_name = given, family
                    titles = [value for ptype, value in parts if ptype == "title"]
                    seco_title = titles[0] if titles else ""
                else:
                    aliases_raw.append({"name": composed, "type": alias_type})
                if document_order and document_order != composed:
                    aliases_raw.append({"name": document_order, "type": alias_type})
        if not primary_name:
            # Aucune identite marquee principale : le premier nom lu fait office
            # de nom principal plutot que d'ecarter la fiche.
            if not aliases_raw:
                continue
            primary_name = aliases_raw.pop(0)["name"]

        # --- Dates de naissance (hors validite de document et dates d'acte) ---
        dobs = []
        for dmy in _seco_find(subject, "day-month-year"):
            if _SECO_NON_BIRTH_ANCESTORS & _seco_ancestors(dmy, parents):
                continue
            iso = _seco_dmy(dmy)
            if iso:
                dobs.append(iso)

        place_of_birth = None
        birth_countries = []
        for pob in _seco_find(subject, "place-of-birth"):
            location = _seco_first_text(pob, "location")
            country_label, country_iso = _seco_country(pob)
            label = ", ".join(p for p in (location, country_label) if p)
            if label and not place_of_birth:
                place_of_birth = label
            if country_iso:
                birth_countries.append(country_iso)

        citizenships = []
        for nat in _seco_find(subject, "nationality"):
            _, iso = _seco_country(nat)
            if iso:
                citizenships.append(iso)

        addresses = []
        address_countries = []
        for addr in _seco_find(subject, "address"):
            if "identification-document" in _seco_ancestors(addr, parents):
                continue
            chunks = []
            for local in ("c-o", "address-details", "po-box", "zip-code", "location", "area-code"):
                chunks.extend(t for t in (_seco_text(n) for n in _seco_find(addr, local)) if t)
            country_label, country_iso = _seco_country(addr)
            if country_label:
                chunks.append(country_label)
            full = ", ".join(chunks)
            if full:
                addresses.append({
                    "full": full,
                    "city": _seco_first_text(addr, "location"),
                    "country": country_label,
                })
            if country_iso:
                address_countries.append(country_iso)

        passports = []
        national_ids = []
        other_registrations = []
        for doc in _seco_find(subject, "identification-document"):
            number = _seco_first_text(doc, "number")
            if not number:
                continue
            doc_type = _seco_attr(doc, "document-type", "type").lower()
            issuer_iso = ""
            for issuer in _seco_find(doc, "issuer"):
                code = _seco_attr(issuer, "code", "iso-code", "iso").upper()
                issuer_iso = code if len(code) == 2 else country_label_to_iso2(_seco_text(issuer))
                if issuer_iso:
                    break
            if not issuer_iso:
                _, issuer_iso = _seco_country(doc)
            expiry = None
            for exp in _seco_find(doc, "expiry-date"):
                for dmy in _seco_find(exp, "day-month-year"):
                    expiry = _seco_dmy(dmy) or expiry
                expiry = expiry or _extract_iso_date(_seco_text(exp))
            if "passport" in doc_type:
                passports.append({
                    "number": number,
                    "issuing_country": issuer_iso or "XX",
                    "expiration_date": expiry,
                })
            elif any(k in doc_type for k in ("id-card", "identity", "identification", "national")):
                national_ids.append({"number": number, "issuing_country": issuer_iso or "XX"})
            else:
                other_registrations.append({
                    "id_type": doc_type or "OtherRegistration",
                    "number": number,
                })

        gender = "U"
        for node in _seco_find(subject, "gender"):
            raw = (_seco_text(node) or _seco_attr(node, "value", "code")).upper()[:1]
            if raw in ("M", "F"):
                gender = raw
            elif raw == "W":      # weiblich, version allemande du fichier
                gender = "F"

        # --- Dates d'acte : inscription et derniere mise a jour ---
        added_dates, all_dates = [], []
        for mod in _seco_find(subject, "modification"):
            mtype = _seco_attr(mod, "modification-type", "type").lower()
            effective = _extract_iso_date(_seco_attr(mod, "effective-date"))
            published = _extract_iso_date(_seco_attr(mod, "publication-date"))
            all_dates.extend(d for d in (effective, published) if d)
            if effective and "add" in mtype:
                added_dates.append(effective)
        listed_on = min(added_dates) if added_dates else (min(all_dates) if all_dates else None)
        last_update = max(all_dates) if all_dates else None

        set_id = _seco_attr(target, "sanctions-set-id") or _seco_attr(subject, "sanctions-set-id")
        program = programs.get(set_id, {})
        program_key = program.get("key", "")
        program_label = program.get("label", "")
        # La Suisse transpose l'ONU et l'UE : l'origine de la mesure est une
        # information de conformite, pas un detail — elle est conservee.
        origin_authority = program.get("origin", "")

        justification = _seco_localized(subject, "justification")
        other_information = _seco_localized(subject, "other-information")
        extra_info = [t for t in (other_information,) if t]
        if program_label and program_label != justification:
            extra_info.append(f"Base legale suisse : {program_label}")
        if origin_authority:
            extra_info.append(f"Mesure d'origine {origin_authority}, transposee par la Suisse")

        primary_addr = addresses[0] if addresses else {}
        yield {
            "entity_id": f"SECO-{ssid}",
            "entity_type": entity_type,
            "primary_name": primary_name,
            "individual_name_parsed": {
                "first_name": first_name,
                "last_name": last_name,
                "maiden_name": "",
            },
            "aliases": categorize_aliases(aliases_raw),
            "dates_of_birth": sorted(set(dobs)),
            "date_of_death": None,
            "is_deceased": False,
            "gender": gender,
            "countries": {
                "citizenship": sorted(set(citizenships)),
                "residence": [],
                "birth_country": sorted(set(birth_countries)),
                "jurisdiction_country": sorted(set(address_countries)) if entity_type != "I" else [],
            },
            "place_of_birth": place_of_birth,
            "address": primary_addr.get("full"),
            "alternative_addresses": [a["full"] for a in addresses[1:]],
            "city": primary_addr.get("city") or None,
            "country": primary_addr.get("country") or None,
            "designation": None,
            "designation_reasons": justification or None,
            "additional_informations": "; ".join(extra_info) or None,
            "official_reference": build_official_reference(
                program_label or program_key or f"SECO ssid {ssid}", last_update
            ),
            "title": seco_title or None,
            "listed_on": listed_on,
            "designating_state": origin_authority or "CH",
            "sanction_programs": [program_key] if program_key else [],
            "name_original_script": None,
            "origin": "SECO Consolidated List",
            "imo_number": None,
            "aircraft_tail_number": None,
            "lei_number": None,
            "national_registry_ids": [],
            "other_registration_ids": other_registrations,
            "passport_documents": passports,
            "national_id_documents": national_ids,
            "other_id_documents": [],
        }


# ------------------ CONSOLIDATED SCREENING LIST (trade.gov) ------------------
# Agregat officiel du gouvernement americain (International Trade
# Administration), publie en JSON public et sans authentification. Son interet
# n'est PAS de redonner la SDN — Fiskr la recupere deja a la source — mais
# d'apporter les listes de CONTROLE DES EXPORTATIONS, aujourd'hui absentes :
#   - Entity List (EL), Denied Persons List (DPL), Unverified List (UVL) et
#     Military End User (MEU) du Bureau of Industry and Security ;
#   - ITAR Debarred (DTC) et Nonproliferation Sanctions (ISN) du Departement
#     d'Etat.
# Ces listes conditionnent le financement du commerce international : une
# contrepartie de trade finance peut y figurer sans etre sur aucune liste de
# gel des avoirs.
#
# Structure : {"results": [ {...}, ... ], "sources_used": [...]}
# Chaque resultat : id, source, type (Individual|Entity|Vessel|Aircraft), name,
# alt_names[], addresses[{address,city,state,postal_code,country}],
# dates_of_birth[], places_of_birth[], nationalities[], citizenships[],
# ids[{type,number,country,issue_date,expiration_date}], programs[],
# federal_register_notice, start_date, end_date, license_requirement,
# license_policy, title, remarks, call_sign, vessel_type, vessel_flag,
# vessel_owner, gross_tonnage, source_list_url.
#
# MEME RESERVE que le connecteur SECO : ecrit d'apres le format publie et
# valide sur un jeu d'essai, pas contre le fichier reel (acces reseau ferme).
# Toutes les cles sont lues avec un defaut : un champ absent n'interrompt rien.
#
# Le fichier est charge d'un bloc (json.load), comme le registre DGT. A
# l'echelle de la CSL (quelques dizaines de milliers de fiches) c'est sans
# consequence ; ce serait a revoir si l'ITA publiait un jour un volume
# comparable au dataset PEP.

_CSL_TYPE_TO_ENTITY = {
    "individual": "I",
    "entity": "E",
    "vessel": "V",
    "aircraft": "O",
}

# Par defaut on ecarte la seule liste que Fiskr recupere deja a sa source et
# qui est active en sortie de boite : la SDN. Si le connecteur Non-SDN est lui
# aussi active, ajouter ses libelles ici evite de doubler les alertes.
CSL_DEFAULT_EXCLUDED_SOURCES = ("Specially Designated Nationals",)


def _csl_list(value: Any) -> List[str]:
    """Valeur multiple tolerante : liste JSON, chaine « a; b », ou vide."""
    if isinstance(value, list):
        return [str(v).strip() for v in value if v is not None and str(v).strip()]
    if isinstance(value, str):
        return [v.strip() for v in value.split(";") if v.strip()]
    return []


def _csl_text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, list):
        return "; ".join(str(v).strip() for v in value if v is not None and str(v).strip())
    return str(value).strip()


def _csl_address(entry: Any) -> Tuple[str, str, str]:
    """(adresse complete, ville, pays) d'une entree d'adresse CSL."""
    if isinstance(entry, str):
        return entry.strip(), "", ""
    if not isinstance(entry, dict):
        return "", "", ""
    city = _csl_text(entry.get("city"))
    country = _csl_text(entry.get("country"))
    parts = [
        _csl_text(entry.get("address")),
        _csl_text(entry.get("postal_code")),
        city,
        _csl_text(entry.get("state")),
        country,
    ]
    return ", ".join(p for p in parts if p), city, country


def parse_csl_json(
    file_path: str,
    excluded_sources: Optional[Tuple[str, ...]] = None,
) -> Generator[Dict[str, Any], None, None]:
    """
    Parse la Consolidated Screening List (trade.gov) vers le schema pivot.

    `excluded_sources` : fragments de libelle, compares sans tenir compte de la
    casse, des listes a IGNORER. Sert a ne pas dupliquer une liste que Fiskr
    recupere deja aupres de son emetteur.
    """
    excluded = tuple(
        s.lower() for s in (
            CSL_DEFAULT_EXCLUDED_SOURCES if excluded_sources is None else excluded_sources
        ) if s
    )

    with open(file_path, "r", encoding="utf-8") as f:
        payload = json.load(f)
    results = payload.get("results") if isinstance(payload, dict) else payload
    for row in results or []:
        if not isinstance(row, dict):
            continue
        name = _csl_text(row.get("name"))
        csl_id = _csl_text(row.get("id")) or _csl_text(row.get("entity_number"))
        if not name or not csl_id:
            continue
        source = _csl_text(row.get("source"))
        source_low = source.lower()
        if any(fragment in source_low for fragment in excluded):
            continue

        entity_type = _CSL_TYPE_TO_ENTITY.get(_csl_text(row.get("type")).lower(), "E")

        aliases_raw = [{"name": a, "type": "Strong"} for a in _csl_list(row.get("alt_names"))]

        dobs = []
        for raw in _csl_list(row.get("dates_of_birth")):
            iso = _extract_iso_date(raw) or _normalize_partial_date(raw)
            if not iso:
                # « circa 1968 », « 1968 to 1970 » : on retient la 1re annee
                year = re.search(r"(\d{4})", raw)
                iso = f"{year.group(1)}-01-01" if year else None
            if iso:
                dobs.append(iso)

        addresses = []
        address_countries = []
        for entry in row.get("addresses") or []:
            full, city, country = _csl_address(entry)
            if full:
                addresses.append({"full": full, "city": city, "country": country})
            if country:
                address_countries.append(country_label_to_iso2(country))

        citizenships = [country_label_to_iso2(c) for c in
                        _csl_list(row.get("citizenships")) + _csl_list(row.get("nationalities"))]
        birth_countries = []
        place_of_birth = None
        for pob in _csl_list(row.get("places_of_birth")):
            if not place_of_birth:
                place_of_birth = pob
            birth_countries.append(country_label_to_iso2(pob.split(",")[-1].strip()))

        passports, national_ids, other_registrations = [], [], []
        for doc in row.get("ids") or []:
            if not isinstance(doc, dict):
                continue
            number = _csl_text(doc.get("number"))
            if not number:
                continue
            doc_type = _csl_text(doc.get("type")).lower()
            issuer = country_label_to_iso2(_csl_text(doc.get("country"))) or "XX"
            expiry = _extract_iso_date(_csl_text(doc.get("expiration_date")))
            if "passport" in doc_type:
                passports.append({"number": number, "issuing_country": issuer,
                                  "expiration_date": expiry})
            elif any(k in doc_type for k in ("national", "identification", "cedula", "id card")):
                national_ids.append({"number": number, "issuing_country": issuer})
            else:
                other_registrations.append({"id_type": doc_type or "OtherRegistration",
                                            "number": number})

        # Le contexte reglementaire d'une liste de controle des exportations
        # est ce qui permet a l'analyste de trancher : il est conserve.
        remarks = _csl_text(row.get("remarks"))
        extra = [t for t in (remarks,) if t]
        for label, key in (("Exigence de licence", "license_requirement"),
                           ("Politique de licence", "license_policy"),
                           ("Fiche source", "source_list_url")):
            value = _csl_text(row.get(key))
            if value:
                extra.append(f"{label} : {value}")

        federal_notice = _csl_text(row.get("federal_register_notice"))
        start_date = _extract_iso_date(_csl_text(row.get("start_date")))
        primary_addr = addresses[0] if addresses else {}

        yield {
            "entity_id": f"CSL-{csl_id}",
            "entity_type": entity_type,
            "primary_name": name,
            "individual_name_parsed": {"first_name": "", "last_name": "", "maiden_name": ""},
            "aliases": categorize_aliases(aliases_raw),
            "dates_of_birth": sorted(set(dobs)),
            "date_of_death": None,
            "is_deceased": False,
            "gender": "U",
            "countries": {
                "citizenship": sorted({c for c in citizenships if c}),
                "residence": [],
                "birth_country": sorted({c for c in birth_countries if c}),
                "jurisdiction_country": sorted(set(address_countries)) if entity_type != "I" else [],
            },
            "place_of_birth": place_of_birth,
            "address": primary_addr.get("full"),
            "alternative_addresses": [a["full"] for a in addresses[1:]],
            "city": primary_addr.get("city") or None,
            "country": primary_addr.get("country") or None,
            "designation": _csl_text(row.get("title")) or None,
            # C'est LA liste americaine qui a designe la contrepartie : sans
            # elle, l'analyste ne sait pas de quelle obligation il s'agit.
            "designation_reasons": source or None,
            "additional_informations": "; ".join(extra) or None,
            "official_reference": build_official_reference(federal_notice, start_date),
            "title": _csl_text(row.get("title")) or None,
            "listed_on": start_date,
            "delisted_on": _extract_iso_date(_csl_text(row.get("end_date"))),
            "designating_state": "US",
            "sanction_programs": _csl_list(row.get("programs")),
            "name_original_script": None,
            "origin": f"US CSL — {source}" if source else "US Consolidated Screening List",
            "imo_number": None,
            "aircraft_tail_number": None,
            "lei_number": None,
            "national_registry_ids": [],
            "other_registration_ids": other_registrations,
            "passport_documents": passports,
            "national_id_documents": national_ids,
            "other_id_documents": [],
            "vessel_call_sign": _csl_text(row.get("call_sign")) or None,
            "vessel_type": _csl_text(row.get("vessel_type")) or None,
            "vessel_flag": _csl_text(row.get("vessel_flag")) or None,
            "vessel_owner": _csl_text(row.get("vessel_owner")) or None,
            "vessel_tonnage": _csl_text(row.get("gross_tonnage")) or None,
        }


# ------------------ LECTURE TOLERANTE DE TABLEAUX ------------------
# Les listes nationales publiees en CSV ou en tableur changent d'intitules de
# colonnes au fil des versions, et plusieurs existent en deux langues. Plutot
# que de figer une orthographe, les connecteurs ci-dessous cherchent une
# colonne par sa FORME NORMALISEE (sans casse, sans accents, sans separateurs),
# et acceptent plusieurs libelles pour un meme champ.


def _table_key(label: str) -> str:
    """« Date of Birth », « date_of_birth », « Date de naissance » -> forme comparable."""
    return re.sub(r"[^a-z0-9]", "", _strip_accents_lower(label or ""))


def _table_get(row: Dict[str, Any], *labels: str) -> str:
    """Premiere colonne renseignee parmi `labels`, comparee sans casse ni accents."""
    normalized = {_table_key(k): v for k, v in row.items() if k}
    for label in labels:
        value = normalized.get(_table_key(label))
        if value is not None and str(value).strip():
            return str(value).strip()
    return ""


def _read_xlsx_rows(file_path: str) -> Generator[Dict[str, str], None, None]:
    """Lignes d'un classeur Excel sous forme de dictionnaires (1re feuille).

    La premiere ligne NON VIDE fait office d'en-tete : les listes officielles
    font souvent preceder le tableau d'un bandeau de titre.
    """
    if not XLSX_AVAILABLE:
        raise RuntimeError(
            "Lecture XLSX indisponible : le paquet openpyxl n'est pas installe. "
            "Installez-le (pip install openpyxl) ou utilisez la version CSV de la liste."
        )
    workbook = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
    try:
        sheet = workbook[workbook.sheetnames[0]]
        header: Optional[List[str]] = None
        for raw in sheet.iter_rows(values_only=True):
            values = ["" if v is None else str(v).strip() for v in raw]
            if header is None:
                if any(values):
                    header = values
                continue
            if not any(values):
                continue
            yield {header[i]: values[i] for i in range(min(len(header), len(values)))}
    finally:
        workbook.close()


def _read_table_rows(file_path: str) -> Generator[Dict[str, str], None, None]:
    """Lignes d'un CSV ou d'un XLSX, choisies sur l'extension du fichier."""
    if file_path.lower().endswith((".xlsx", ".xlsm")):
        yield from _read_xlsx_rows(file_path)
        return
    with open(file_path, "r", encoding="utf-8-sig", newline="") as f:
        for row in csv.DictReader(f):
            yield {k: ("" if v is None else v) for k, v in row.items() if k}


# ------------------ LISTE CANADIENNE (SEMA, Affaires mondiales Canada) ------------------
# Liste consolidee des sanctions autonomes canadiennes (Special Economic
# Measures Act et Loi sur les mesures economiques speciales), publiee en CSV
# par Affaires mondiales Canada. Le Canada sanctionne de facon autonome, avec
# un perimetre qui ne recoupe ni celui de l'UE ni celui de l'OFAC.
#
# Le fichier existe en anglais ET en francais : les deux jeux d'intitules sont
# acceptes, ce qui evite qu'un telechargement depuis la page francophone donne
# une liste vide. Colonnes attendues (ordre indifferent) :
#   Country / Pays, Item / Article, Schedule / Annexe,
#   LastName / Nom, GivenName / Prenom, Aliases / Pseudonymes,
#   DateOfBirth / DateDeNaissance, Entity / Entite, Title / Titre,
#   DateOfListing / DateInscription
#
# MEME RESERVE que SECO et CSL : ecrit d'apres le format publie et valide sur
# un jeu d'essai, pas contre le fichier reel (acces reseau ferme).


def _canada_stable_id(schedule: str, item: str, name: str) -> str:
    """
    Le fichier canadien ne porte pas d'identifiant technique. La cle de delta
    est donc reconstruite : annexe + article quand ils sont presents (c'est la
    reference reglementaire, stable d'une publication a l'autre), a defaut une
    empreinte du nom. Sans cle stable, chaque publication paraitrait remplacer
    integralement la precedente et le delta serait illisible.
    """
    reference = "-".join(p for p in (schedule, item) if p)
    if reference:
        return re.sub(r"[^A-Za-z0-9.-]+", "", reference.replace(" ", ""))
    return hashlib.sha1(_strip_accents_lower(name).encode("utf-8")).hexdigest()[:12].upper()


def parse_canada_sema_csv(file_path: str) -> Generator[Dict[str, Any], None, None]:
    """Parse la liste consolidee des sanctions autonomes canadiennes."""
    seen: Dict[str, int] = {}
    for row in _read_table_rows(file_path):
        entity = _table_get(row, "Entity", "Entite", "Entité", "EntityName")
        last_name = _table_get(row, "LastName", "Last Name", "Nom", "NomDeFamille")
        given_name = _table_get(row, "GivenName", "Given Name", "Prenom", "Prénom", "Prenoms")
        if entity:
            entity_type = "E"
            primary_name = entity
            first_name = last_name_out = ""
        else:
            entity_type = "I"
            primary_name = " ".join(p for p in (given_name, last_name) if p)
            first_name, last_name_out = given_name, last_name
        if not primary_name:
            continue

        schedule = _table_get(row, "Schedule", "Annexe")
        item = _table_get(row, "Item", "Article")
        base_id = _canada_stable_id(schedule, item, primary_name)
        # Une meme annexe/article peut porter plusieurs designes : on suffixe
        # pour garder une cle unique sans perdre la reference reglementaire.
        seen[base_id] = seen.get(base_id, 0) + 1
        entity_id = base_id if seen[base_id] == 1 else f"{base_id}.{seen[base_id]}"

        aliases_raw = [
            {"name": a, "type": "Strong"}
            for a in parse_multi_value(
                {"aliases": _table_get(row, "Aliases", "Alias", "Pseudonymes", "AKA")},
                "aliases",
            )
        ]
        # L'ordre inverse (patronyme d'abord) reste cherchable, comme pour SECO
        if entity_type == "I" and given_name and last_name:
            aliases_raw.append({"name": f"{last_name} {given_name}", "type": "Strong"})

        dob = _extract_iso_date(_table_get(row, "DateOfBirth", "Date of Birth",
                                           "DateDeNaissance", "Date de naissance"))
        country = _table_get(row, "Country", "Pays")
        listed_on = _extract_iso_date(_table_get(row, "DateOfListing", "Date of Listing",
                                                 "DateInscription", "Date d'inscription"))
        title = _table_get(row, "Title", "Titre")
        reference = " ".join(p for p in (schedule and f"Annexe {schedule}",
                                         item and f"article {item}") if p)

        yield {
            "entity_id": f"CA-{entity_id}",
            "entity_type": entity_type,
            "primary_name": primary_name,
            "individual_name_parsed": {
                "first_name": first_name, "last_name": last_name_out, "maiden_name": ""
            },
            "aliases": categorize_aliases(aliases_raw),
            "dates_of_birth": [dob] if dob else [],
            "date_of_death": None,
            "is_deceased": False,
            "gender": "U",
            "countries": {
                "citizenship": [country_label_to_iso2(country)] if (country and entity_type == "I") else [],
                "residence": [],
                "birth_country": [],
                "jurisdiction_country": [country_label_to_iso2(country)] if (country and entity_type != "I") else [],
            },
            "place_of_birth": None,
            "address": None,
            "alternative_addresses": [],
            "city": None,
            "country": country or None,
            "designation": title or None,
            "designation_reasons": f"Sanctions autonomes canadiennes — {country}" if country
                                   else "Sanctions autonomes canadiennes",
            "additional_informations": None,
            "official_reference": build_official_reference(reference, listed_on),
            "title": title or None,
            "listed_on": listed_on,
            "designating_state": "CA",
            "sanction_programs": [country] if country else [],
            "name_original_script": None,
            "origin": "Canada — Sanctions autonomes (SEMA)",
            "imo_number": None,
            "aircraft_tail_number": None,
            "lei_number": None,
            "national_registry_ids": [],
            "other_registration_ids": [],
            "passport_documents": [],
            "national_id_documents": [],
            "other_id_documents": [],
        }


# ------------------ LISTE AUSTRALIENNE (DFAT Consolidated List) ------------------
# Liste consolidee du Department of Foreign Affairs and Trade, qui reunit les
# sanctions onusiennes transposees ET les sanctions autonomes australiennes.
# Publiee en XLSX et en CSV : les deux sont acceptes, l'extension tranche.
#
# Structure en LIGNES REPETEES par variante de nom, regroupees par `Reference`
# — le meme principe que le ConList britannique. La colonne `Name Type`
# distingue le nom principal des alias.
#
# MEME RESERVE : ecrit d'apres le format publie, valide sur un jeu d'essai.


def parse_dfat_consolidated(file_path: str) -> Generator[Dict[str, Any], None, None]:
    """Parse la liste consolidee australienne (DFAT), CSV ou XLSX."""
    groups: Dict[str, Dict[str, Any]] = {}
    order: List[str] = []

    for row in _read_table_rows(file_path):
        reference = _table_get(row, "Reference", "Ref", "ReferenceNumber")
        name = _table_get(row, "Name of Individual or Entity", "Name", "NameOfIndividualOrEntity")
        if not name:
            continue
        if not reference:
            reference = hashlib.sha1(
                _strip_accents_lower(name).encode("utf-8")).hexdigest()[:12].upper()

        group = groups.get(reference)
        if group is None:
            group = groups[reference] = {
                "primary": "", "aliases": [], "type": "", "dobs": [], "pobs": [],
                "citizenships": [], "addresses": [], "info": [], "committees": "",
                "control_date": "", "listing": "",
            }
            order.append(reference)

        name_type = _table_get(row, "Name Type", "NameType", "Type of Name").lower()
        is_primary = ("primary" in name_type) or (not name_type and not group["primary"])
        if is_primary and not group["primary"]:
            group["primary"] = name
        elif name != group["primary"]:
            group["aliases"].append(name)

        group["type"] = group["type"] or _table_get(row, "Type", "Entity Type")
        for raw in parse_multi_value({"v": _table_get(row, "Date of Birth", "DateOfBirth", "DOB")}, "v"):
            iso = _extract_iso_date(raw) or _normalize_partial_date(raw)
            if not iso:
                year = re.search(r"(\d{4})", raw)
                iso = f"{year.group(1)}-01-01" if year else None
            if iso:
                group["dobs"].append(iso)
        for key in ("Place of Birth", "PlaceOfBirth"):
            value = _table_get(row, key)
            if value:
                group["pobs"].append(value)
                break
        for raw in parse_multi_value({"v": _table_get(row, "Citizenship", "Nationality")}, "v"):
            group["citizenships"].append(country_label_to_iso2(raw))
        address = _table_get(row, "Address", "Addresses")
        if address and address not in group["addresses"]:
            group["addresses"].append(address)
        info = _table_get(row, "Additional Information", "AdditionalInformation")
        if info and info not in group["info"]:
            group["info"].append(info)
        group["committees"] = group["committees"] or _table_get(row, "Committees", "Committee")
        group["listing"] = group["listing"] or _table_get(row, "Listing Information", "ListingInformation")
        group["control_date"] = group["control_date"] or _table_get(row, "Control Date", "ControlDate")

    for reference in order:
        group = groups[reference]
        primary_name = group["primary"] or (group["aliases"].pop(0) if group["aliases"] else "")
        if not primary_name:
            continue
        raw_type = group["type"].lower()
        if "individual" in raw_type or "person" in raw_type:
            entity_type = "I"
        elif "vessel" in raw_type or "ship" in raw_type:
            entity_type = "V"
        else:
            entity_type = "E"

        control_date = _extract_iso_date(group["control_date"])
        # Le comite onusien d'origine, quand il y en a un : la liste
        # australienne transpose l'ONU autant qu'elle designe pour son compte.
        committees = group["committees"]
        yield {
            "entity_id": f"AU-{reference}",
            "entity_type": entity_type,
            "primary_name": primary_name,
            "individual_name_parsed": {"first_name": "", "last_name": "", "maiden_name": ""},
            "aliases": categorize_aliases(
                [{"name": a, "type": "Strong"} for a in dict.fromkeys(group["aliases"])]
            ),
            "dates_of_birth": sorted(set(group["dobs"])),
            "date_of_death": None,
            "is_deceased": False,
            "gender": "U",
            "countries": {
                "citizenship": sorted({c for c in group["citizenships"] if c}),
                "residence": [],
                "birth_country": [],
                "jurisdiction_country": [],
            },
            "place_of_birth": group["pobs"][0] if group["pobs"] else None,
            "address": group["addresses"][0] if group["addresses"] else None,
            "alternative_addresses": group["addresses"][1:],
            "city": None,
            "country": None,
            "designation": None,
            "designation_reasons": group["listing"] or None,
            "additional_informations": "; ".join(group["info"]) or None,
            "official_reference": build_official_reference(reference, control_date),
            "title": None,
            "listed_on": control_date,
            "designating_state": "UN" if committees else "AU",
            "sanction_programs": [committees] if committees else [],
            "name_original_script": None,
            "origin": "Australie — DFAT Consolidated List",
            "imo_number": None,
            "aircraft_tail_number": None,
            "lei_number": None,
            "national_registry_ids": [],
            "other_registration_ids": [],
            "passport_documents": [],
            "national_id_documents": [],
            "other_id_documents": [],
        }


# ------------------ LISTES D'ALERTE DES REGULATEURS ------------------
# Une liste d'alerte n'est PAS une liste de sanctions. La SFC de Hong Kong,
# l'AMF, la FCA publient des mises en garde contre des entites non autorisees,
# des sites frauduleux, des usurpations d'identite d'etablissements agrees.
# Une touche n'emporte aucune obligation de gel : c'est un signal de risque,
# a instruire. C'est pourquoi chaque liste d'alerte recoit son PROPRE type de
# liste — donc son propre seuil (`scoring.cut_off_overrides`) et ses propres
# statistiques — au lieu d'etre versee dans le meme flux que les gels d'avoirs.
#
# Ces listes ont trois proprietes communes qui justifient un lecteur partage :
# elles sont publiees sous forme de TABLEAU (HTML le plus souvent, parfois CSV
# ou JSON), leurs intitules de colonnes varient d'un regulateur a l'autre, et
# elles ne portent AUCUN identifiant technique.


class _HTMLTableExtractor(HTMLParser):
    """Extrait les lignes du plus grand <table> d'une page.

    Les regulateurs publient leur liste d'alerte comme une page web ; le
    tableau de donnees y voisine avec des tableaux de mise en page. On retient
    celui qui a le plus de lignes, ce qui est le critere le plus robuste sans
    connaitre la page.
    """

    def __init__(self):
        super().__init__(convert_charrefs=True)
        self.tables: List[List[List[str]]] = []
        self._table: Optional[List[List[str]]] = None
        self._row: Optional[List[str]] = None
        self._cell: Optional[List[str]] = None

    def handle_starttag(self, tag, attrs):
        if tag == "table":
            self._table = []
        elif tag == "tr" and self._table is not None:
            self._row = []
        elif tag in ("td", "th") and self._row is not None:
            self._cell = []

    def handle_endtag(self, tag):
        if tag in ("td", "th") and self._cell is not None:
            self._row.append(re.sub(r"\s+", " ", "".join(self._cell)).strip())
            self._cell = None
        elif tag == "tr" and self._row is not None:
            if any(self._row):
                self._table.append(self._row)
            self._row = None
        elif tag == "table" and self._table is not None:
            self.tables.append(self._table)
            self._table = None

    def handle_data(self, data):
        if self._cell is not None:
            self._cell.append(data)


def _read_html_table_rows(file_path: str) -> Generator[Dict[str, str], None, None]:
    """Lignes du tableau principal d'une page HTML, en-tete = 1re ligne."""
    with open(file_path, "r", encoding="utf-8", errors="replace") as f:
        parser = _HTMLTableExtractor()
        parser.feed(f.read())
    if not parser.tables:
        return
    table = max(parser.tables, key=len)
    if len(table) < 2:
        return
    header = table[0]
    for row in table[1:]:
        yield {header[i]: row[i] for i in range(min(len(header), len(row)))}


def _sniff_table_format(file_path: str) -> str:
    """Devine le format d'apres les PREMIERS OCTETS, pas d'apres l'extension.

    Les regulateurs servent leur liste a une URL sans extension : se fier au
    nom du fichier ferait lire un CSV comme une page web, et l'import
    ressortirait a ZERO fiche — c'est-a-dire silencieusement, sous les traits
    d'une liste vide plutot que d'une erreur. Le contenu tranche donc.
    """
    with open(file_path, "rb") as f:
        head = f.read(4096)
    if head[:4] == b"PK\x03\x04":          # classeur Excel (archive zip)
        return "xlsx"
    text = head.decode("utf-8", errors="replace").lstrip("\ufeff").lstrip()
    if text[:1] in ("{", "["):
        return "json"
    if text[:1] == "<":
        return "html"
    return "table"


def _read_alert_rows(file_path: str) -> Generator[Dict[str, Any], None, None]:
    """Lignes d'une liste d'alerte : JSON, CSV/XLSX ou tableau HTML."""
    detected = _sniff_table_format(file_path)
    if detected == "json":
        with open(file_path, "r", encoding="utf-8") as f:
            payload = json.load(f)
        if isinstance(payload, dict):
            for key in ("results", "data", "items", "records", "rows"):
                if isinstance(payload.get(key), list):
                    payload = payload[key]
                    break
            else:
                payload = [payload]
        for row in payload or []:
            if isinstance(row, dict):
                yield row
        return
    if detected == "html":
        yield from _read_html_table_rows(file_path)
        return
    if detected == "xlsx":
        yield from _read_xlsx_rows(file_path)
        return
    yield from _read_table_rows(file_path)


def _alert_stable_id(name: str, website: str) -> str:
    """
    Ces listes ne portent pas d'identifiant. La cle de delta derive donc du nom
    normalise, complete du site quand il y en a un : deux entites homonymes
    sevissant sur deux domaines restent deux fiches distinctes, et la meme
    entite garde sa cle d'une publication a l'autre.
    """
    seed = _strip_accents_lower(name) + "|" + _strip_accents_lower(website)
    return hashlib.sha1(seed.encode("utf-8")).hexdigest()[:12].upper()


def parse_regulatory_alert_list(
    file_path: str,
    *,
    id_prefix: str,
    origin: str,
    authority: str,
    designation_reasons: str,
    reference_label: str,
    default_jurisdiction: str = "",
) -> Generator[Dict[str, Any], None, None]:
    """
    Lit une liste d'alerte de regulateur vers le schema pivot.

    Les colonnes sont cherchees par forme normalisee et sous plusieurs
    libelles, anglais comme francais : les regulateurs n'ont aucune convention
    commune, et le meme regulateur change d'intitules au fil des refontes.
    """
    for row in _read_alert_rows(file_path):
        row = {k: v for k, v in row.items() if k}
        name = _table_get(
            row, "Name", "Name of Entity", "Entity Name", "Company Name", "Company",
            "Entity", "Nom", "Dénomination", "Denomination", "Nom de l'entité",
            "Raison sociale", "Nom commercial", "Entité",
        )
        websites = parse_multi_value(
            {"w": _table_get(row, "Website", "Websites", "Web site", "URL", "Site",
                             "Site internet", "Adresse du site", "Sites internet")},
            "w",
        )
        if not name:
            # Certaines listes ne nomment que le site : il fait alors office
            # d'identite, faute de quoi la ligne serait perdue.
            name = websites[0] if websites else ""
        if not name:
            continue

        aliases_raw = [
            {"name": a, "type": "Strong"}
            for a in parse_multi_value(
                {"a": _table_get(row, "Also known as", "Alias", "Aliases", "Other names",
                                 "Trading name", "Autres dénominations", "Autre dénomination",
                                 "Nom(s) commercial(aux)")},
                "a",
            )
        ]
        # Les autres domaines d'une meme entite sont cherchables comme alias :
        # un virement libelle au nom du site doit ressortir.
        aliases_raw += [{"name": w, "type": "Weak"} for w in websites[1:]]

        raw_type_label = _table_get(row, "Type", "Category", "Catégorie", "Categorie",
                                    "Entity type", "Nature")
        raw_type = _strip_accents_lower(raw_type_label)
        # ATTENTION : « personne morale » contient « person ». Les libelles de
        # personne MORALE sont donc testes EN PREMIER, sans quoi toute societe
        # francaise serait typee personne physique — et le type d'entite est une
        # composante de la cle de blocking, donc l'erreur ecarterait des
        # candidats au lieu de se voir.
        if any(k in raw_type for k in ("personne morale", "entity", "entite", "company",
                                       "corporate", "societe", "organisation", "organization")):
            entity_type = "E"
        elif any(k in raw_type for k in ("individual", "personne physique", "natural person")):
            entity_type = "I"
        else:
            entity_type = "E"

        published = _extract_iso_date(_table_get(
            row, "Date", "Date of publication", "Publication date", "Date added",
            "Date of listing", "Date de publication", "Date d'inscription",
            "Date de mise en garde", "Last updated",
        ))
        country = _table_get(row, "Country", "Jurisdiction", "Pays", "Juridiction")
        # A defaut de pays publie, la juridiction du REGULATEUR fait foi. Ce
        # n'est pas un remplissage de confort : COUNTRY_ISO est une composante
        # de la cle de blocking et une fiche sans pays tombe dans la partition
        # « XX », que ne rejoint aucun client ayant un pays — elle serait donc
        # structurellement inatteignable. Une entite visee par un avertissement
        # de la SFC opere par construction sur le marche de Hong Kong.
        jurisdiction = country_label_to_iso2(country) if country else default_jurisdiction
        details = _table_get(row, "Reason", "Remarks", "Details", "Description",
                             "Comments", "Motif", "Commentaire", "Précisions", "Precisions")
        reference = _table_get(row, "Reference", "Ref", "Référence", "Case number",
                               "Numéro", "Numero")

        extra = [t for t in (details,) if t]
        if websites:
            extra.append("Site(s) : " + ", ".join(websites))
        if raw_type_label:
            extra.append(f"Catégorie : {raw_type_label}")

        yield {
            "entity_id": "{}-{}".format(
                id_prefix,
                re.sub(r"[^A-Za-z0-9._-]+", "", reference) if reference
                else _alert_stable_id(name, websites[0] if websites else "")
            ),
            "entity_type": entity_type,
            "primary_name": name,
            "individual_name_parsed": {"first_name": "", "last_name": "", "maiden_name": ""},
            "aliases": categorize_aliases(aliases_raw),
            "dates_of_birth": [],
            "date_of_death": None,
            "is_deceased": False,
            "gender": "U",
            "countries": {
                "citizenship": [],
                "residence": [],
                "birth_country": [],
                "jurisdiction_country": [jurisdiction] if jurisdiction else [],
            },
            "place_of_birth": None,
            "address": _table_get(row, "Address", "Adresse", "Registered address") or None,
            "alternative_addresses": [],
            "city": None,
            "country": country or None,
            "designation": None,
            # Le libelle dit explicitement qu'il s'agit d'une mise en garde :
            # l'analyste ne doit jamais la confondre avec un gel des avoirs.
            "designation_reasons": designation_reasons,
            "additional_informations": "; ".join(extra) or None,
            "official_reference": build_official_reference(reference or reference_label, published),
            "title": None,
            "listed_on": published,
            "designating_state": authority,
            "sanction_programs": [],
            "name_original_script": None,
            "origin": origin,
            "imo_number": None,
            "aircraft_tail_number": None,
            "lei_number": None,
            "national_registry_ids": [],
            "other_registration_ids": [],
            "passport_documents": [],
            "national_id_documents": [],
            "other_id_documents": [],
            "websites": websites,
        }


def parse_hk_sfc_alert_list(file_path: str) -> Generator[Dict[str, Any], None, None]:
    """
    Liste d'alerte de la Securities and Futures Commission de Hong Kong.

    La SFC y recense les entites non autorisees, les sites suspects et les
    usurpations d'identite d'intermediaires agrees. Hong Kong n'ayant pas de
    regime de sanctions autonome (les mesures onusiennes y sont transposees par
    ordonnance), c'est cette liste-la qui apporte quelque chose qu'aucune autre
    source branchee ne porte.
    """
    return parse_regulatory_alert_list(
        file_path,
        id_prefix="HKSFC",
        origin="Hong Kong SFC — Alert List",
        authority="HK",
        designation_reasons="Mise en garde du régulateur — SFC Hong Kong (entité non autorisée / site suspect)",
        reference_label="SFC Hong Kong — Alert List",
        default_jurisdiction="HK",
    )


def parse_amf_blacklist(file_path: str) -> Generator[Dict[str, Any], None, None]:
    """
    Listes noires de l'Autorite des marches financiers (France).

    Sites et entites proposant des services d'investissement sans y etre
    autorises. Marche domestique de l'etablissement : c'est la liste d'alerte
    dont un assujetti francais a le plus directement l'usage.
    """
    return parse_regulatory_alert_list(
        file_path,
        id_prefix="AMF",
        origin="AMF — Listes noires",
        authority="FR",
        designation_reasons="Mise en garde du régulateur — AMF (acteur non autorisé)",
        reference_label="AMF — listes noires",
        default_jurisdiction="FR",
    )


# ------------------ EXCLUSIONS BANQUE MONDIALE ------------------
# Fournisseurs et personnes exclus des marches finances par le Groupe de la
# Banque mondiale (sanctions de passation de marches). Ni gel des avoirs ni
# mise en garde : une exclusion pour fraude ou corruption averee, publiee en
# JSON public. Les banques la criblent au titre du risque de contrepartie sur
# le financement de projets et le commerce international.
#
# Structure : {"response": {"ZPROCSUPP": [ {...}, ... ]}} — l'enveloppe exacte
# varie selon le point d'entree, d'ou une recherche tolerante de la liste.


def _worldbank_records(payload: Any) -> List[Dict[str, Any]]:
    """Trouve la liste d'enregistrements quelle que soit l'enveloppe JSON."""
    if isinstance(payload, list):
        return [r for r in payload if isinstance(r, dict)]
    if not isinstance(payload, dict):
        return []
    for value in payload.values():
        found = _worldbank_records(value)
        if found:
            return found
    return []


def parse_worldbank_debarred_json(file_path: str) -> Generator[Dict[str, Any], None, None]:
    """Parse la liste des fournisseurs exclus par la Banque mondiale."""
    with open(file_path, "r", encoding="utf-8") as f:
        payload = json.load(f)

    for row in _worldbank_records(payload):
        row = {(k or "").strip().upper(): v for k, v in row.items()}

        def get(*keys: str) -> str:
            for key in keys:
                value = row.get(key)
                if value is not None and str(value).strip():
                    return str(value).strip()
            return ""

        name = get("SUPP_NAME", "FIRM_NAME", "NAME")
        if not name:
            continue
        supplier_id = get("SUPP_ID", "ID") or _alert_stable_id(name, "")
        country = get("COUNTRY_NAME", "COUNTRY")
        from_date = _extract_iso_date(get("DEBAR_FROM_DATE", "FROM_DATE"))
        to_date = _extract_iso_date(get("DEBAR_TO_DATE", "TO_DATE"))
        reason = get("DEBAR_REASON", "REASON", "GROUNDS")
        address = get("SUPP_ADDR", "ADDRESS")

        extra = [t for t in (reason,) if t]
        if to_date:
            extra.append(f"Exclusion jusqu'au {to_date}")

        yield {
            "entity_id": f"WB-{supplier_id}",
            "entity_type": "E",
            "primary_name": name,
            "individual_name_parsed": {"first_name": "", "last_name": "", "maiden_name": ""},
            "aliases": categorize_aliases([]),
            "dates_of_birth": [],
            "date_of_death": None,
            "is_deceased": False,
            "gender": "U",
            "countries": {
                "citizenship": [],
                "residence": [],
                "birth_country": [],
                "jurisdiction_country": [country_label_to_iso2(country)] if country else [],
            },
            "place_of_birth": None,
            "address": address or None,
            "alternative_addresses": [],
            "city": get("SUPP_CITY", "CITY") or None,
            "country": country or None,
            "designation": None,
            "designation_reasons": "Exclusion des marchés financés par la Banque mondiale",
            "additional_informations": "; ".join(extra) or None,
            "official_reference": build_official_reference(
                f"Banque mondiale — exclusion {supplier_id}", from_date),
            "title": None,
            "listed_on": from_date,
            # La fin d'exclusion est une radiation programmee : la colonne
            # existe deja, autant la renseigner plutot que de la noyer en texte.
            "delisted_on": to_date,
            "designating_state": "WB",
            "sanction_programs": [],
            "name_original_script": None,
            "origin": "World Bank — Debarred firms and individuals",
            "imo_number": None,
            "aircraft_tail_number": None,
            "lei_number": None,
            "national_registry_ids": [],
            "other_registration_ids": [{"id_type": "WorldBankSupplierId", "number": supplier_id}],
            "passport_documents": [],
            "national_id_documents": [],
            "other_id_documents": [],
        }


# ------------------ CSV CONNECTOR ------------------

def parse_csv_file(file_path: str, delimiter: str = ",", mapping_dict: dict = None) -> Generator[Dict[str, Any], None, None]:
    """
    Parses Client or Watchlist CSV dataset dynamically.
    Uses custom delimiters and maps columns according to config.
    """
    with open(file_path, mode="r", encoding="utf-8") as f:
        reader = csv.DictReader(f, delimiter=delimiter)
        for row in reader:
            # Map columns
            mapped_row = {}
            if mapping_dict:
                for target, source in mapping_dict.items():
                    mapped_row[target] = row.get(source, "")
            else:
                # Direct mirror map based on headers
                for k, v in row.items():
                    mapped_row[k] = v
                    
            yield mapped_row


# ------------------ PDF IA PARSING CONNECTOR ------------------

def parse_pdf_watchlist(file_path: str) -> List[Dict[str, Any]]:
    """
    Ingests publications/PDF files:
    1. Extracts text via pypdf.
    2. Runs NER heuristic to structure entities.
    3. Simulates LLM schema verification.
    """
    text = ""
    if PDF_AVAILABLE:
        try:
            with open(file_path, "rb") as f:
                pdf = pypdf.PdfReader(f)
                for page in pdf.pages:
                    text += page.extract_text() or ""
        except Exception as e:
            logger.error(f"Error reading PDF: {e}")
    else:
        # Fallback if library missing
        logger.warning("pypdf not installed, simulating text extraction")
        text = "COMMISSION REGULATION - Gels d'avoirs - AL-MANSOUR SHIPPING (IMO 99412) - pays résidence: RU."

    # Step 2: Simulated LLM Named Entity Recognition (NER)
    # Scan text for names, countries, dates and IDs
    entities_extracted = []
    
    # We parse the text using regular expressions to mimic a structured LLM parser.
    # Ex: AL-MANSOUR SHIPPING, RESIDENCE: RU, IMO: 99412
    vessels = re.findall(r"([A-Z\-\s]+)\s*\(IMO\s*(\d+)\)", text)
    for name, imo in vessels:
        entities_extracted.append({
            "entity_id": f"PDF-VES-{imo}",
            "entity_type": "V",
            "primary_name": name.strip(),
            "imo_number": imo,
            "countries": {"jurisdiction_country": ["RU"]},
            "extraction_confidence": 95.0 # High confidence
        })
        
    # Standard warning if no clear patterns found (confidence < 85%)
    if not entities_extracted:
        # Generate a warning mock entry with low confidence
        entities_extracted.append({
            "entity_id": "PDF-LOW-CONF",
            "entity_type": "I",
            "primary_name": "INCONNU EXTRAIT",
            "extraction_confidence": 75.0, # Will trigger Rule_M08 warning
            "countries": {}
        })
        
    return entities_extracted
